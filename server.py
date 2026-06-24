from __future__ import annotations

import datetime as dt
import concurrent.futures
import csv
import html
import io
import json
import os
import posixpath
import re
import shutil
import subprocess
import sys
import tempfile
import time as _time
import urllib.parse
import urllib.request
import zipfile
from copy import copy as copy_style, deepcopy
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font


BASE_DIR = Path(__file__).resolve().parent
REPORT_DIR = BASE_DIR / "reports"
SOURCE_DIR = BASE_DIR / "sources"
TEMPLATE_DIR = BASE_DIR / "templates"
TEMPLATE_PATH = TEMPLATE_DIR / "同業送件明細樣板.xlsx"
SFB_PAGE = "https://www.sfb.gov.tw/ch/home.jsp?id=1016&parentpath=0%2C6%2C52"
ENABLE_ONLINE_MOPS_LOOKUP = os.environ.get("ENABLE_ONLINE_MOPS_LOOKUP", "1").lower() in {"1", "true", "yes", "on"}
MAX_MOPS_WORKERS = int(os.environ.get("MAX_MOPS_WORKERS", "4"))

COMPANY_TYPES = ("上市", "上櫃")
CASE_KEYWORDS = (
    "現金增資",
    "轉換公司債",
    "交換公司債",
    "存託憑證",
)

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", NS_MAIN)
ET.register_namespace("r", NS_REL)


def ensure_dirs() -> None:
    REPORT_DIR.mkdir(exist_ok=True)
    SOURCE_DIR.mkdir(exist_ok=True)
    TEMPLATE_DIR.mkdir(exist_ok=True)


def roc_year(date: dt.date) -> int:
    return date.year - 1911


def current_friday(today: dt.date | None = None) -> dt.date:
    today = today or dt.date.today()
    monday = today - dt.timedelta(days=today.weekday())
    return monday + dt.timedelta(days=4)


def default_week(today: dt.date | None = None) -> tuple[dt.date, dt.date]:
    friday = current_friday(today)
    return friday - dt.timedelta(days=4), friday


def report_window(end: dt.date) -> tuple[dt.date, dt.date]:
    return end - dt.timedelta(days=4), end


def roc_date(date: dt.date, sep: str = "/") -> str:
    return f"{roc_year(date):03d}{sep}{date.month:02d}{sep}{date.day:02d}"


def compact_roc_date(date: dt.date) -> str:
    return f"{roc_year(date):03d}{date.month:02d}{date.day:02d}"


def parse_date(value: str) -> dt.date | None:
    text = str(value or "").strip()
    if not text:
        return None
    text = text.replace("民國", "").replace("年", "/").replace("月", "/").replace("日", "")
    text = re.sub(r"\.0$", "", text)
    digits = re.sub(r"\D", "", text)
    try:
        if len(digits) == 7:
            year = int(digits[:3]) + 1911
            return dt.date(year, int(digits[3:5]), int(digits[5:7]))
        if len(digits) == 8:
            return dt.date(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))
        if "/" in text:
            parts = [p for p in re.split(r"\D+", text) if p]
            if len(parts) >= 3:
                year = int(parts[0])
                if year < 1911:
                    year += 1911
                return dt.date(year, int(parts[1]), int(parts[2]))
    except ValueError:
        return None
    return None


def normalize_header(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).replace("　", "")


def xml_text(elem: ET.Element) -> str:
    return "".join(elem.itertext())


def col_to_number(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + ord(ch.upper()) - 64
    return n


def number_to_col(n: int) -> str:
    out = ""
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


def cell_col(ref: str) -> str:
    match = re.match(r"([A-Z]+)", ref or "")
    return match.group(1) if match else ""


def read_xlsx(path: Path) -> list[dict[str, object]]:
    ns = {"m": NS_MAIN, "r": NS_REL, "pr": NS_PKG_REL}
    rows_by_sheet: list[dict[str, object]] = []
    with zipfile.ZipFile(path) as zf:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            shared = [xml_text(si) for si in root.findall("m:si", ns)]

        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_map = {}
        for rel in rels.findall("pr:Relationship", ns):
            target = rel.attrib.get("Target", "")
            if not target.startswith("/"):
                target = "xl/" + target
            rel_map[rel.attrib["Id"]] = posixpath.normpath(target.lstrip("/"))

        sheets = workbook.findall("m:sheets/m:sheet", ns)
        for sheet in sheets:
            name = sheet.attrib.get("name", "工作表")
            rid = sheet.attrib.get(f"{{{NS_REL}}}id")
            target = rel_map.get(rid or "", "")
            if not target:
                continue
            ws = ET.fromstring(zf.read(target))
            parsed_rows = []
            for row in ws.findall(".//m:sheetData/m:row", ns):
                row_data: dict[str, str] = {}
                for cell in row.findall("m:c", ns):
                    ref = cell.attrib.get("r", "")
                    col = cell_col(ref)
                    cell_type = cell.attrib.get("t")
                    value = ""
                    if cell_type == "inlineStr":
                        inline = cell.find("m:is", ns)
                        value = xml_text(inline) if inline is not None else ""
                    else:
                        raw = cell.find("m:v", ns)
                        value = raw.text if raw is not None and raw.text is not None else ""
                        if cell_type == "s" and value:
                            idx = int(float(value))
                            value = shared[idx] if 0 <= idx < len(shared) else value
                    if col:
                        row_data[col] = value.strip()
                if row_data:
                    parsed_rows.append(row_data)
            rows_by_sheet.append({"name": name, "rows": parsed_rows})
    return rows_by_sheet


def find_header(rows: list[dict[str, str]]) -> tuple[int, dict[str, str]]:
    aliases = {
        "證券代號": ("證券代號", "代號"),
        "公司型態": ("公司型態",),
        "結案類型": ("結案類型", "辦理情形"),
        "公司名稱": ("公司名稱", "公司"),
        "承銷商": ("承銷商",),
        "案件類別": ("案件類別",),
        "金額": ("金額",),
        "幣別": ("幣別",),
        "發行價格": ("發行價格",),
        "收文日期": ("收文日期",),
        "自動補正日期": ("自動補正日期",),
        "停止生效日期": ("停止生效日期",),
        "解除生效日期": ("解除生效日期",),
        "生效日期": ("生效日期",),
        "廢止撤銷日期": ("廢止/撤銷日期", "廢止撤銷日期"),
        "自行撤回日期": ("自行撤回日期",),
        "退件日期": ("退件日期",),
        "案件性質": ("案件性質",),
        "承銷方式": ("承銷方式",),
    }
    best_idx = -1
    best_map: dict[str, str] = {}
    best_score = 0
    for i, row in enumerate(rows[:20]):
        normalized = {col: normalize_header(value) for col, value in row.items()}
        found: dict[str, str] = {}
        for canonical, names in aliases.items():
            for col, value in normalized.items():
                if any(name == value or name in value for name in names):
                    found[canonical] = col
                    break
        score = sum(1 for key in ("公司名稱", "承銷商", "案件類別") if key in found) + len(found)
        if score > best_score:
            best_idx = i
            best_map = found
            best_score = score
    if best_score < 5:
        raise ValueError("找不到可辨識的欄位列，請確認來源檔是證期局年度申報案件 Excel。")
    return best_idx, best_map


def row_value(row: dict[str, str], header: dict[str, str], name: str) -> str:
    return row.get(header.get(name, ""), "").strip()


def case_code(case_type: str) -> str:
    text = normalize_header(case_type)
    if "存託憑證" in text:  # 涵蓋「海外存託憑證」與「存託憑證(海外)」兩種寫法
        return "GDR"
    if "交換公司債" in text:
        return "EB"
    if "海外" in text and "轉換公司債" in text:
        return "ECB"
    if "轉換公司債" in text:
        return "CB"
    if "現金增資" in text:
        return "CI"
    return "其他"


SECURITY_SHORT_NAMES = {
    "1295": "生合",
    "5291": "邑昇",
    "8442": "威宏-KY",
    "6465": "威潤",
    "6223": "旺矽",
    "3372": "典範",
    "1717": "長興",
    "9935": "慶豐富",
    "8299": "群聯",
}

BOND_SHORT_NAMES = {
    ("1295", "CB", "400000000"): "生合一",
    ("5291", "CB", "200000000"): "邑昇二",
    ("8442", "CB", "400000000"): "威宏三-KY",
    ("8442", "CB", "200000000"): "威宏四-KY",
    ("6223", "CB", "5000000000"): "旺矽六",
    ("1717", "CB", "2000000000"): "長興二",
    ("9935", "CB", "300000000"): "慶豐富四",
    ("8299", "ECB", "800000000"): "群聯海外一",
}

MOPS_TIMEOUT_SECONDS = 2

# Hard-coded seed entries (bootstraps the cache on first deploy)
MOPS_ENRICHMENTS: dict[tuple[str, str, str], dict[str, str]] = {
    ("1295", "CB", "400000000"): {"display": "生合一", "purpose": "償還銀行借款"},
    ("1295", "CI", "31250000"): {"display": "生合", "purpose": "償還銀行借款"},
    ("5291", "CB", "200000000"): {"display": "邑昇二", "purpose": "償還銀行借款"},
    ("5291", "CI", "20000000"): {"display": "邑昇", "purpose": "償還銀行借款"},
    ("8442", "CB", "400000000"): {"display": "威宏三-KY", "purpose": "償還銀行借款；充實營運資金；支應新客戶開發需求"},
    ("8442", "CB", "200000000"): {"display": "威宏四-KY", "purpose": "償還銀行借款；充實營運資金；支應新客戶開發需求"},
    ("8442", "CI", "40000000"): {"display": "威宏-KY", "purpose": "償還銀行借款；充實營運資金；支應新客戶開發需求"},
    ("6465", "CI", "120000000"): {"display": "威潤", "purpose": "充實營運資金"},
    ("6223", "CB", "5000000000"): {"display": "旺矽六", "purpose": "充實營運資金"},
    ("3372", "CI", "600000000"): {"display": "典範", "purpose": "購置機器設備"},
}

BOND_CACHE_PATH = BASE_DIR / "data" / "bond_cache.json"
_BOND_CACHE_LOCK = __import__("threading").Lock()


def _load_bond_cache() -> None:
    """Merge persisted cache into MOPS_ENRICHMENTS at startup."""
    if not BOND_CACHE_PATH.exists():
        return
    try:
        import json
        raw = json.loads(BOND_CACHE_PATH.read_text(encoding="utf-8"))
        for k, v in raw.items():
            parts = k.split("\x00")
            if len(parts) == 3:
                key = (parts[0], parts[1], parts[2])
                MOPS_ENRICHMENTS.setdefault(key, v)
    except Exception:
        pass


def _save_bond_cache() -> None:
    """Flush MOPS_ENRICHMENTS to disk (only CB/ECB/EB entries worth persisting)."""
    try:
        import json
        BOND_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        keep = {"\x00".join(k): v for k, v in MOPS_ENRICHMENTS.items()
                if k[1] in ("CB", "ECB", "EB") and v.get("display")}
        with _BOND_CACHE_LOCK:
            BOND_CACHE_PATH.write_text(json.dumps(keep, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


_load_bond_cache()


def public_fetch_text(url: str, data: dict[str, str] | None = None, timeout: int = MOPS_TIMEOUT_SECONDS) -> str:
    import requests as _requests
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://mopsov.twse.com.tw/mops/web/index",
    }
    try:
        if data is not None:
            resp = _requests.post(url, data=data, headers=headers, timeout=timeout)
        else:
            resp = _requests.get(url, headers=headers, timeout=timeout)
        resp.raise_for_status()
        raw = resp.content
        charset = resp.encoding or "utf-8"
    except Exception:
        if data is not None:
            raise
        raw = fetch_with_windows(url, attempt_timeout=max(5, min(timeout + 2, 8)))
        charset = "utf-8"
    for candidate in (charset, "utf-8", "big5", "cp950"):
        try:
            return raw.decode(candidate)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def html_to_text(value: str) -> str:
    text = re.sub(r"<script[\s\S]*?</script>", " ", value, flags=re.I)
    text = re.sub(r"<style[\s\S]*?</style>", " ", text, flags=re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    text = html.unescape(text)
    return re.sub(r"\s+", " ", text).strip()


_COMPANY_SHORT_CACHE: dict[str, str] = {}
_BULK_COMPANY_LOADED = False
import threading as _threading_early
_BULK_COMPANY_LOCK = _threading_early.Lock()


def _bulk_load_company_names() -> None:
    """Download full company lists once and populate _COMPANY_SHORT_CACHE for all codes."""
    global _BULK_COMPANY_LOADED
    with _BULK_COMPANY_LOCK:
        if _BULK_COMPANY_LOADED:
            return
        _BULK_COMPANY_LOADED = True
    sources = (
        ("json", "https://openapi.twse.com.tw/v1/opendata/t187ap03_L"),
        ("json", "https://openapi.twse.com.tw/v1/opendata/t187ap03_O"),
    )
    for fmt, url in sources:
        try:
            text = public_fetch_text(url, timeout=5)
            if fmt == "json":
                rows = json.loads(text)
                if not isinstance(rows, list):
                    continue
                for row in rows:
                    if not isinstance(row, dict):
                        continue
                    code = str(row.get("公司代號") or row.get("有價證券代號") or row.get("Code") or "").strip()
                    name = str(row.get("公司簡稱") or row.get("有價證券名稱") or row.get("Name") or "").strip()
                    if code and name and code not in _COMPANY_SHORT_CACHE:
                        _COMPANY_SHORT_CACHE[code] = name
            else:
                reader = csv.DictReader(io.StringIO(text))
                for row in reader:
                    code = str(row.get("公司代號") or row.get("有價證券代號") or row.get("證券代號") or "").strip()
                    name = str(row.get("公司簡稱") or row.get("有價證券名稱") or row.get("公司名稱") or "").strip()
                    if code and name and code not in _COMPANY_SHORT_CACHE:
                        _COMPANY_SHORT_CACHE[code] = name
        except Exception:
            continue


def public_company_short_name(security_code: str) -> str:
    if not _BULK_COMPANY_LOADED:
        _bulk_load_company_names()
    if security_code in _COMPANY_SHORT_CACHE:
        return _COMPANY_SHORT_CACHE[security_code]
    # Fallback: per-stock fast API
    for market in ("tse", "otc"):
        url = f"https://mis.twse.com.tw/stock/api/getStockInfo.jsp?ex_ch={market}_{security_code}.tw&json=1&delay=0"
        try:
            data = json.loads(public_fetch_text(url, timeout=2))
        except Exception:
            continue
        rows = data.get("msgArray") if isinstance(data, dict) else []
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            if str(row.get("c") or "").strip() != security_code:
                continue
            name = str(row.get("n") or "").strip()
            if name:
                _COMPANY_SHORT_CACHE[security_code] = name
                return name
    mops_name = mops_company_short_name(security_code)
    if mops_name:
        _COMPANY_SHORT_CACHE[security_code] = mops_name
        return mops_name
    _COMPANY_SHORT_CACHE[security_code] = ""
    return ""


def mops_company_short_name(security_code: str) -> str:
    if not security_code:
        return ""
    params = {
        "encodeURIComponent": "1",
        "step": "1",
        "firstin": "1",
        "off": "1",
        "keyword4": "",
        "code1": "",
        "TYPEK2": "",
        "checkbtn": "",
        "queryName": "co_id",
        "inpuType": "co_id",
        "TYPEK": "all",
        "co_id": security_code,
    }
    text = mops_query_text(("/mops/web/ajax_t05st03", "/mops/web/t05st03"), params)
    compact = normalize_header(html_to_text(text))
    if security_code not in compact:
        return ""
    patterns = (
        rf"{re.escape(security_code)}(?:公司簡稱|簡稱)[：:]?([\u4e00-\u9fffA-Za-z0-9-]{{2,16}})",
        r"(?:公司簡稱|簡稱)[：:]?([\u4e00-\u9fffA-Za-z0-9-]{2,16})",
        rf"{re.escape(security_code)}([\u4e00-\u9fffA-Za-z0-9-]{{2,16}})(?:上市|上櫃|興櫃|公開發行)",
    )
    for pattern in patterns:
        for match in re.finditer(pattern, compact):
            candidate = match.group(1).strip()
            if is_plausible_security_short(candidate):
                return candidate
    return ""


def is_plausible_security_short(value: str) -> bool:
    if not value or len(value) > 12:
        return False
    bad_words = ("公司", "股價", "股票", "法人", "新聞", "即時", "基本", "財報", "營收", "董事", "代號", "名稱")
    return not any(word in value for word in bad_words)


def stock_display_name(record: dict[str, str]) -> str:
    short_name = public_company_short_name(record.get("證券代號", ""))
    if short_name:
        return normalize_stock_short_for_bond(short_name)
    return normalize_stock_short_for_bond(display_name_for_record(record) or record.get("公司名稱", ""))


def mops_query_text(paths: tuple[str, ...], params: dict[str, str]) -> str:
    hosts = ("https://mopsov.twse.com.tw", "https://mops.twse.com.tw")
    urls = [host + path for host in hosts for path in paths]

    def _fetch_one(url: str) -> str:
        try:
            return public_fetch_text(url + "?" + urllib.parse.urlencode(params), timeout=MOPS_TIMEOUT_SECONDS)
        except Exception:
            try:
                return public_fetch_text(url, data=params, timeout=MOPS_TIMEOUT_SECONDS)
            except Exception:
                return ""

    with concurrent.futures.ThreadPoolExecutor(max_workers=len(urls)) as ex:
        results = list(ex.map(_fetch_one, urls))
    return "\n".join(r for r in results if r)


def mops_follow_detail_links(text: str) -> str:
    links: list[str] = []
    for raw in re.findall(r"""(?:href|action)=["']([^"']*t05sr01_1[^"']*)["']""", text, flags=re.I):
        link = html.unescape(raw)
        if link.startswith("/"):
            links.append("https://mops.twse.com.tw" + link)
            links.append("https://mopsov.twse.com.tw" + link)
        elif link.startswith("http"):
            links.append(link)
    for raw in re.findall(r"""(?:https?://[^"'\s<>]+|/mops/web/[^"'\s<>]+)""", text, flags=re.I):
        if "t05sr01_1" not in raw:
            continue
        link = html.unescape(raw)
        if link.startswith("/"):
            links.append("https://mops.twse.com.tw" + link)
            links.append("https://mopsov.twse.com.tw" + link)
        else:
            links.append(link)
    detail_texts: list[str] = []
    for link in dict.fromkeys(links):
        try:
            detail_texts.append(public_fetch_text(link, timeout=MOPS_TIMEOUT_SECONDS))
        except Exception:
            continue
    return "\n".join(detail_texts)


def mops_subject_search_text(record: dict[str, str], date_value: dt.date, keyword: str) -> str:
    market_kinds = ["C"]
    if "上市" in record.get("公司型態", ""):
        market_kinds.insert(0, "L")
    elif "上櫃" in record.get("公司型態", ""):
        market_kinds.insert(0, "O")
    texts: list[str] = []
    for kind in dict.fromkeys(market_kinds):
        params = {
            "encodeURIComponent": "1",
            "step": "1",
            "firstin": "true",
            "id": "",
            "key": "",
            "TYPEK": "",
            "Stp": "4",
            "go": "false",
            "co_id": record.get("證券代號", ""),
            "COMPANY_ID": record.get("證券代號", ""),
            "r1": "1",
            "KIND": kind,
            "CODE": "",
            "keyWord": keyword,
            "Condition2": "2",
            "keyWord2": "",
            "year": f"{roc_year(date_value):03d}",
            "month1": str(date_value.month),
            "begin_day": "1",
            "end_day": "31",
            "Orderby": "1",
        }
        query_text = mops_query_text(("/mops/web/ajax_t51sb10", "/mops/web/t51sb10"), params)
        texts.append(query_text)
        texts.append(mops_follow_detail_links(query_text))
    return "\n".join(texts)


_MOPS_LOOKUP_CACHE: dict[tuple, str] = {}


def mops_official_lookup_text(record: dict[str, str], end: dt.date, include_bond: bool = True) -> str:
    received = parse_date(record.get("收文日期", "")) or end
    co_id = record.get("證券代號", "")
    classification = record.get("分類", "")
    cache_key = (co_id, classification, received.year, received.month, end.year, end.month, include_bond)
    cached = _MOPS_LOOKUP_CACHE.get(cache_key)
    if cached is not None:
        return cached
    texts: list[str] = []

    # Primary: announcement list for received month (single call, 4 parallel requests)
    params = {
        "encodeURIComponent": "1",
        "step": "1",
        "firstin": "1",
        "TYPEK": "all",
        "co_id": co_id,
        "year": f"{roc_year(received):03d}",
        "month": f"{received.month:02d}",
    }
    texts.append(mops_query_text(("/mops/web/ajax_t05sr01_1", "/mops/web/t05sr01_1"), params))

    # If received and end are different months, also query end month
    if (roc_year(end), end.month) != (roc_year(received), received.month):
        end_params = dict(params)
        end_params["year"] = f"{roc_year(end):03d}"
        end_params["month"] = f"{end.month:02d}"
        texts.append(mops_query_text(("/mops/web/ajax_t05sr01_1", "/mops/web/t05sr01_1"), end_params))

    texts.append(open_data_major_announcement_text(record))
    result = "\n".join(text for text in texts if text)
    _MOPS_LOOKUP_CACHE[cache_key] = result
    return result


def parse_bond_short_from_text(text: str, record: dict[str, str], company_short: str = "") -> str:
    if not text:
        return ""
    compact = html_to_text(text)
    security_code = record.get("證券代號", "")
    online_or_source = company_short or record.get("顯示名稱", "") or record.get("公司名稱", "")
    fallback = SECURITY_SHORT_NAMES.get(security_code, "")
    bases = {
        online_or_source,
        online_or_source.replace("-KY", ""),
        record.get("公司名稱", "").replace("F-", "").replace("-KY", ""),
        fallback,
        fallback.replace("-KY", ""),
    }
    suffix = r"(?:海外)?[一二三四五六七八九十百\d]+(?:-KY)?"
    candidates: list[str] = []
    for item in bases:
        item = item.strip()
        if not item:
            continue
        for match in re.finditer(re.escape(item) + suffix, compact):
            value = match.group(0)
            if len(value) <= len(item):
                continue
            candidates.append(value)
    if record.get("分類") == "ECB":
        overseas = [value for value in candidates if "海外" in value or "ECB" in value.upper()]
        if overseas:
            return normalize_bond_product_name(overseas[0], record)
    return normalize_bond_product_name(candidates[0], record) if candidates else ""


def mops_bond_short_name(record: dict[str, str], end: dt.date, company_short: str = "") -> str:
    if record.get("分類") not in ("CB", "ECB", "EB"):
        return ""
    text = mops_official_lookup_text(record, end, include_bond=True)
    name = parse_bond_short_from_text(text, record, company_short)
    if name:
        return name
    return parse_bond_short_from_announcement(text, record, company_short)


def parse_purpose_from_text(text: str) -> str:
    compact = normalize_header(html_to_text(text))
    if not compact:
        return ""
    purposes: list[str] = []
    checks = (
        ("償還銀行借款", ("償還銀行借款", "償還金融機構借款")),
        ("充實營運資金", ("充實營運資金", "充實營運週轉金", "營運資金")),
        ("購置機器設備", ("購置機器設備", "購置設備")),
        ("擴建廠房", ("擴建廠房", "興建廠房")),
        ("轉投資國內外事業", ("轉投資", "投資國內外")),
        ("償還前次", ("償還前次", "償還第")),
    )
    for label, keywords in checks:
        if any(keyword in compact for keyword in keywords):
            purposes.append(label)
    return "；".join(dict.fromkeys(purposes))


def funding_purpose_from_open_text(record: dict[str, str], text: str) -> str:
    matched = matching_record_text(text, record)
    return parse_purpose_from_text(matched)


def amount_search_text(amount: str) -> str:
    digits = re.sub(r"\D", "", amount or "")
    if not digits:
        return ""
    try:
        value = int(digits)
    except ValueError:
        return digits
    if value >= 100000000 and value % 100000000 == 0:
        return f"{value // 100000000}億元"
    if value >= 100000000:
        return f"{value / 100000000:g}億元"
    if value >= 10000 and value % 10000 == 0:
        return f"{value // 10000}萬元"
    return digits


def open_data_major_announcement_text(record: dict[str, str]) -> str:
    urls = (
        "https://openapi.twse.com.tw/v1/opendata/t187ap04_L",
        "https://www.tpex.org.tw/openapi/v1/mopsfin_t187ap04_O",
        "https://www.tpex.org.tw/openapi/v1/t187ap04_O",
        "https://dts.twse.com.tw/opendata/t187ap04_L.csv",
        "https://dts.twse.com.tw/opendata/t187ap04_O.csv",
        "https://dts.twse.com.tw/opendata/t187ap04_P.csv",
        "http://dts.twse.com.tw/opendata/t187ap04_L.csv",
        "http://dts.twse.com.tw/opendata/t187ap04_O.csv",
        "http://dts.twse.com.tw/opendata/t187ap04_P.csv",
        "https://mopsfin.twse.com.tw/opendata/t187ap04_L.csv",
        "https://mopsfin.twse.com.tw/opendata/t187ap04_O.csv",
        "https://mopsfin.twse.com.tw/opendata/t187ap04_P.csv",
    )
    all_rows = _load_open_data_rows(urls)
    security_code = record.get("證券代號", "")
    tokens = record_identity_tokens(record)
    texts: list[str] = []
    for code, joined, joined_norm in all_rows:
        if code and code != security_code:
            continue
        if security_code in joined or any(token in joined_norm for token in tokens):
            texts.append(joined)
    return "\n".join(texts)


_OPEN_DATA_ROWS: list[tuple[str, str, str]] | None = None
_OPEN_DATA_LOCK = _threading_early.Lock()


def _load_open_data_rows(urls: tuple[str, ...]) -> list[tuple[str, str, str]]:
    """Download the market-wide open-data files once; cache parsed rows for all records."""
    global _OPEN_DATA_ROWS
    if _OPEN_DATA_ROWS is not None:
        return _OPEN_DATA_ROWS
    with _OPEN_DATA_LOCK:
        if _OPEN_DATA_ROWS is not None:
            return _OPEN_DATA_ROWS
        seen_codes: set[str] = set()
        parsed: list[tuple[str, str, str]] = []
        for url in urls:
            try:
                raw = public_fetch_text(url, timeout=4)
            except Exception:
                continue
            if not raw.strip():
                continue
            rows: list[dict[str, str]] = []
            try:
                data = json.loads(raw)
                if isinstance(data, list):
                    rows = [row for row in data if isinstance(row, dict)]
            except Exception:
                try:
                    rows = list(csv.DictReader(io.StringIO(raw)))
                except Exception:
                    rows = []
            if not rows:
                continue
            for row in rows:
                code = str(row.get("公司代號") or row.get("證券代號") or row.get("公司代號/股票代號") or row.get("Code") or "").strip()
                joined = " ".join(str(value) for value in row.values() if value is not None)
                dedup = code + "|" + joined
                if dedup in seen_codes:
                    continue
                seen_codes.add(dedup)
                parsed.append((code, joined, normalize_header(joined)))
        _OPEN_DATA_ROWS = parsed
        return _OPEN_DATA_ROWS


def text_matches_record(text: str, record: dict[str, str]) -> bool:
    return bool(matching_record_text(text, record))


def record_identity_tokens(record: dict[str, str]) -> list[str]:
    identities = [
        record.get("證券代號", ""),
        record.get("顯示名稱", ""),
        record.get("公司名稱", ""),
        display_name_for_record(record),
    ]
    tokens: list[str] = []
    for item in identities:
        token = normalize_header(item).replace("F-", "").replace("-KY", "")
        if token and token not in tokens:
            tokens.append(token)
    return tokens


def record_amount_tokens(record: dict[str, str]) -> set[str]:
    amount = re.sub(r"\D", "", record.get("金額", ""))
    if not amount:
        return set()
    amount_tokens = {amount}
    try:
        value = int(amount)
        if value % 1000 == 0:
            amount_tokens.add(str(value // 1000))
        if value % 1000000 == 0:
            amount_tokens.add(str(value // 1000000))
    except ValueError:
        pass
    return {token for token in amount_tokens if token}


def record_case_tokens(record: dict[str, str]) -> list[str]:
    code = record.get("分類", "")
    case_type = normalize_header(record.get("案件類別", ""))
    if code == "CI":
        return ["現金增資"]
    if code == "CB":
        return ["轉換公司債"]
    if code == "ECB":
        return ["海外轉換公司債", "轉換公司債"]
    if code == "EB":
        return ["交換公司債"]
    if code == "GDR":
        return ["海外存託憑證", "存託憑證"]
    return [case_type] if case_type else []


def matching_record_text(text: str, record: dict[str, str]) -> str:
    compact = normalize_header(html_to_text(text))
    if not compact:
        return ""
    identities = record_identity_tokens(record)
    if not identities:
        return ""
    identity_positions = [
        match.start()
        for token in identities
        for match in re.finditer(re.escape(token), compact)
        if token
    ]
    if not identity_positions:
        return ""
    amount_tokens = record_amount_tokens(record)
    if not amount_tokens:
        return compact if identity_positions else ""

    matched: list[str] = []
    for token in sorted(amount_tokens, key=len, reverse=True):
        for match in re.finditer(re.escape(token), compact):
            pos = match.start()
            if any(abs(pos - identity_pos) <= 1400 for identity_pos in identity_positions):
                start = max(0, pos - 1800)
                end = min(len(compact), pos + 1800)
                matched.append(compact[start:end])
    if matched:
        return " ".join(dict.fromkeys(matched))
    case_tokens = [token for token in record_case_tokens(record) if token]
    for token in case_tokens:
        for match in re.finditer(re.escape(token), compact):
            pos = match.start()
            if any(abs(pos - identity_pos) <= 1800 for identity_pos in identity_positions):
                start = max(0, pos - 2200)
                end = min(len(compact), pos + 2200)
                matched.append(compact[start:end])
    return " ".join(dict.fromkeys(matched))


def mops_major_announcement_text(record: dict[str, str], end: dt.date) -> str:
    received = parse_date(record.get("收文日期", "")) or end
    dates = [received]
    if received.month != end.month or received.year != end.year:
        dates.append(end)
    texts: list[str] = []
    for date in dates:
        typek = "sii" if "上市" in record.get("公司型態", "") else ("otc" if "上櫃" in record.get("公司型態", "") else "all")
        params = {
            "encodeURIComponent": "1",
            "step": "1",
            "firstin": "1",
            "off": "1",
            "queryName": "COMPANY_ID",
            "inpuType": "co_id",
            "TYPEK": typek,
            "co_id": record.get("證券代號", ""),
            "year": f"{roc_year(date):03d}",
            "month": f"{date.month:02d}",
            "b_date": "1",
            "e_date": "31",
        }
        texts.append(mops_query_text(("/mops/web/ajax_t05st01", "/mops/web/t05st01", "/mops/web/ajax_t05sr01_1", "/mops/web/t05sr01_1"), params))
        if typek != "all":
            params_all = dict(params)
            params_all["TYPEK"] = "all"
            texts.append(mops_query_text(("/mops/web/ajax_t05st01", "/mops/web/t05st01"), params_all))
    return "\n".join(texts)


def chinese_ordinal_to_short(value: str) -> str:
    text = normalize_header(value)
    if text.isdigit():
        number = int(text)
        digits = ["零", "一", "二", "三", "四", "五", "六", "七", "八", "九", "十"]
        return digits[number] if 0 <= number <= 10 else text
    return text


CHINESE_ORDINALS = ["零", "一", "二", "三", "四", "五", "六", "七", "八", "九", "十", "十一", "十二", "十三", "十四", "十五", "十六", "十七", "十八", "十九", "二十"]


def ordinal_to_int(value: str) -> int | None:
    text = chinese_ordinal_to_short(value)
    if text.isdigit():
        return int(text)
    if text in CHINESE_ORDINALS:
        return CHINESE_ORDINALS.index(text)
    if text.startswith("十") and len(text) == 2 and text[1] in CHINESE_ORDINALS:
        return 10 + CHINESE_ORDINALS.index(text[1])
    if len(text) == 2 and text[0] in CHINESE_ORDINALS and text[1] == "十":
        return CHINESE_ORDINALS.index(text[0]) * 10
    if len(text) == 3 and text[0] in CHINESE_ORDINALS and text[1] == "十" and text[2] in CHINESE_ORDINALS:
        return CHINESE_ORDINALS.index(text[0]) * 10 + CHINESE_ORDINALS.index(text[2])
    return None


def int_to_chinese_ordinal(value: int) -> str:
    if 0 <= value < len(CHINESE_ORDINALS):
        return CHINESE_ORDINALS[value]
    tens, ones = divmod(value, 10)
    if ones == 0:
        return f"{CHINESE_ORDINALS[tens]}十"
    return f"{CHINESE_ORDINALS[tens]}十{CHINESE_ORDINALS[ones]}"


def normalize_stock_short_for_bond(value: str) -> str:
    base = normalize_header(value).replace("F-", "")
    if base.startswith("MOPS待確認："):
        base = base.removeprefix("MOPS待確認：")
    for suffix in ("股份有限公司", "有限公司", "公司"):
        if base.endswith(suffix) and len(base) > len(suffix) + 1:
            base = base[: -len(suffix)]
    if base.endswith("-KY"):
        ky = "-KY"
        base = base[:-3]
    else:
        ky = ""
    suffix_rules = (
        ("國際科技", ""),
        ("國際", ""),
        ("精密", ""),
        ("精機", ""),
        ("電子", ""),
        ("達科技", "達科"),
        ("科技", ""),
    )
    for suffix, replacement in suffix_rules:
        if base.endswith(suffix) and len(base) >= len(suffix) + 1:
            base = base[: -len(suffix)] + replacement
            break
    return base + ky


def bond_name_with_ordinal(base: str, ordinal: str, record: dict[str, str]) -> str:
    stock = normalize_stock_short_for_bond(base)
    if not stock:
        return ""
    if record.get("分類") == "ECB":
        # ECB convention: arabic numeral with "ECB" prefix, e.g. 華邦電ECB4
        num = ordinal_to_int(ordinal)
        arabic = str(num) if num is not None else ordinal
        if stock.endswith("-KY"):
            return f"{stock[:-3]}ECB{arabic}-KY"
        return f"{stock}ECB{arabic}"
    if stock.endswith("-KY"):
        return f"{stock[:-3]}{ordinal}-KY"
    return f"{stock}{ordinal}"


def normalize_bond_product_name(value: str, record: dict[str, str]) -> str:
    text = normalize_header(value).replace("F-", "")
    if not text:
        return ""
    match = re.match(r"(.+?)(海外)?([一二三四五六七八九十百\d]+)(-KY)?$", text)
    if match:
        base = match.group(1)
        overseas = match.group(2) or ""
        ordinal = chinese_ordinal_to_short(match.group(3))
        suffix = match.group(4) or ""
        if suffix and not base.endswith("-KY"):
            base += "-KY"
        return bond_name_with_ordinal(base, overseas + ordinal, record)
    return text


def is_bond_product_name(value: str) -> bool:
    text = normalize_header(value)
    if not text or len(text) > 18:
        return False
    if any(word in text for word in ("公告", "公司", "證券", "科技股份", "待確認")):
        return False
    return bool(re.search(r"(海外)?[一二三四五六七八九十百\d]+(?:-KY)?$", text))


def bond_product_ordinal(value: str) -> str:
    text = normalize_header(value).removeprefix("MOPS待確認：")
    match = re.search(r"(?:海外)?([一二三四五六七八九十百\d]+)(?:-KY)?$", text)
    return chinese_ordinal_to_short(match.group(1)) if match else ""


def bond_ordinals_from_text(text: str) -> list[str]:
    compact = normalize_header(html_to_text(text))
    ordinals: list[str] = []
    patterns = (
        r"第([一二三四五六七八九十百\d]+)次及第([一二三四五六七八九十百\d]+)次",
        r"第([一二三四五六七八九十百\d]+)次",
        r"國內([一二三四五六七八九十百\d]+)次",
        r"海外([一二三四五六七八九十百\d]+)次",
    )
    for pattern in patterns:
        for match in re.finditer(pattern, compact):
            for group in match.groups():
                if not group:
                    continue
                ordinal = chinese_ordinal_to_short(group)
                if ordinal not in ordinals:
                    ordinals.append(ordinal)
    return ordinals


def bond_lookup_required(record: dict[str, str], focus_keys: set[str] | None) -> bool:
    return (
        record.get("分類") in ("CB", "ECB", "EB")
        and (focus_keys is None or record_key(record) in focus_keys)
        and not is_bond_product_name(record.get("顯示名稱", ""))
    )


def resolve_missing_bond_names(records: list[dict[str, str]], end: dt.date, focus_keys: set[str] | None, warnings: list[str]) -> None:
    groups: dict[str, list[dict[str, str]]] = {}
    for record in records:
        if record.get("分類") not in ("CB", "ECB", "EB") or (focus_keys is not None and record_key(record) not in focus_keys):
            continue
        group_key = "|".join((record.get("證券代號", ""), record.get("公司名稱", ""), record.get("分類", "")))
        groups.setdefault(group_key, []).append(record)
    for group in groups.values():
        if all(is_bond_product_name(record.get("顯示名稱", "")) for record in group):
            continue
        mops_text = mops_official_lookup_text(group[0], end, include_bond=True)
        explicit = parse_bond_short_from_announcement(mops_text, group[0], public_company_short_name(group[0].get("證券代號", "")))
        missing = [record for record in group if not is_bond_product_name(record.get("顯示名稱", ""))]
        if len(group) == 1 and len(missing) == 1 and is_bond_product_name(explicit):
            missing[0]["顯示名稱"] = explicit
            continue
        ordinals = bond_ordinals_from_text(mops_text)
        known_ordinals = [bond_product_ordinal(record.get("顯示名稱", "")) for record in group if is_bond_product_name(record.get("顯示名稱", ""))]
        for known in known_ordinals:
            if known and known not in ordinals:
                ordinals.append(known)
        ordinals = complete_ordinals(ordinals, len(group))
        company_short = public_company_short_name(group[0].get("證券代號", "")) or group[0].get("顯示名稱", "") or group[0].get("公司名稱", "")
        if ordinals:
            for record, ordinal in zip(group, ordinals):
                record["顯示名稱"] = bond_name_with_ordinal(company_short, ordinal, record)
        else:
            for record in group:
                record["顯示名稱"] = stock_display_name(record)


_CN_DIGITS = {
    "零": 0, "一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9,
    "壹": 1, "貳": 2, "參": 3, "叄": 3, "肆": 4, "伍": 5, "陸": 6, "柒": 7, "捌": 8, "玖": 9, "兩": 2,
}


def _parse_yi(text: str) -> int:
    """解析億位數字（阿拉伯或中文大寫），如 '15'->15、'伍'->5、'貳拾'->20、'壹拾伍'->15。"""
    text = text.strip()
    if text.isdigit():
        return int(text)
    if "拾" in text or "十" in text:
        parts = re.split(r"[拾十]", text)
        tens = _CN_DIGITS.get(parts[0], 1) if parts[0] else 1
        ones = _CN_DIGITS.get(parts[1], 0) if len(parts) > 1 and parts[1] else 0
        return tens * 10 + ones
    return _CN_DIGITS.get(text, 0)


def _ordinal_amounts(detail: str) -> dict[str, int]:
    """從公告 detail『發行總額』段落抓 {序號: 發行金額(元)}。
    僅在各次金額分列時有效；若為『第三、四次...合計』則抓不到（回空）。"""
    amounts: dict[str, int] = {}
    for match in re.finditer(
        r"第([一二三四五六七八九十百\d]+)次[^第]*?新台幣([零壹貳參叄肆伍陸柒捌玖拾佰兩一二三四五六七八九十\d]+)億",
        detail,
    ):
        ordinal = chinese_ordinal_to_short(match.group(1))
        yi = _parse_yi(match.group(2))
        if ordinal and yi and ordinal not in amounts:
            amounts[ordinal] = yi * 100000000
    return amounts


def reassign_multi_bond_ordinals(records: list[dict[str, str]], end: dt.date, focus_keys: set[str] | None, warnings: list[str]) -> None:
    """同公司同次申報多檔 CB（常以一則『第X次暨第Y次』公告宣布）時，序號擷取易把
    多筆都標成第一個序號。優先用各檔發行金額與公告金額精準對應；金額對不上的剩餘
    檔再按申報順序補序號並加警示。僅在目前序號有重複時才覆寫（不動已正確的）。"""
    groups: dict[tuple, list[dict[str, str]]] = {}
    for record in records:
        if record.get("分類") not in ("CB", "ECB", "EB"):
            continue
        if focus_keys is not None and record_key(record) not in focus_keys:
            continue
        groups.setdefault((record.get("證券代號", ""), record.get("分類", "")), []).append(record)
    for (code, _cls), group in groups.items():
        if len(group) < 2:
            continue
        current = [bond_product_ordinal(r.get("顯示名稱", "")) for r in group]
        if len(set(current)) >= len(group):
            continue  # 序號已各不相同，視為正確，不動
        ordinals: list[str] = []
        amounts: dict[str, int] = {}
        for title, detail in _company_resolution_announcements(group[0], end):
            compact = re.sub(r"\s+", "", title)
            for raw in re.findall(r"第([一二三四五六七八九十百\d]+)次", compact):
                ordinal = chinese_ordinal_to_short(raw)
                if ordinal not in ordinals:
                    ordinals.append(ordinal)
            amounts.update(_ordinal_amounts(detail))
        ordinals = sorted(set(ordinals), key=lambda o: (ordinal_to_int(o) or 999))
        if len(ordinals) < len(group):
            continue  # 序號不足，保留現狀（避免亂猜）
        short = public_company_short_name(code) or normalize_stock_short_for_bond(group[0].get("公司名稱", ""))
        # 第一輪：用發行金額精準對應序號
        available = ordinals[:]
        assigned: dict[int, str] = {}
        unmatched: list[dict[str, str]] = []
        for record in group:
            amount = int(re.sub(r"\D", "", record.get("金額", "0")) or 0)
            hit = next((o for o in available if amounts.get(o) == amount), None)
            if hit:
                assigned[id(record)] = hit
                available.remove(hit)
            else:
                unmatched.append(record)
        # 第二輪：剩餘檔按申報順序補剩餘序號
        for record, ordinal in zip(unmatched, available):
            assigned[id(record)] = ordinal
        for record in group:
            ordinal = assigned.get(id(record))
            if not ordinal:
                continue
            name = bond_name_with_ordinal(short, ordinal, record)
            if name and is_bond_product_name(name):
                record["顯示名稱"] = name
        if unmatched:
            unmatched_ords = "、".join(assigned[id(r)] for r in unmatched if assigned.get(id(r)))
            warnings.append(
                f"{code} {group[0].get('公司名稱')}：多檔轉換公司債部分序號無法由發行金額精準對應"
                f"（第{unmatched_ords}次依申報順序推定），請人工確認。"
            )


def normalize_weekly_stock_names(records: list[dict[str, str]], focus_keys: set[str] | None) -> None:
    for record in records:
        if focus_keys is not None and record_key(record) not in focus_keys:
            continue
        if record.get("分類") in ("CB", "ECB", "EB"):
            if is_bond_product_name(record.get("顯示名稱", "")):
                continue
        else:
            record["顯示名稱"] = stock_display_name(record)


def require_bond_names(records: list[dict[str, str]], focus_keys: set[str] | None) -> None:
    missing: list[str] = []
    for record in records:
        if record.get("分類") not in ("CB", "ECB", "EB"):
            continue
        if focus_keys is not None and record_key(record) not in focus_keys:
            continue
        if is_bond_product_name(record.get("顯示名稱", "")):
            continue
        missing.append(
            f"{record.get('證券代號', '')} {record.get('公司名稱', '')} "
            f"{record.get('案件類別', '')} {record.get('金額', '')} "
            f"收文{record.get('收文日期', '')}"
        )
    if missing:
        raise ValueError("MOPS 第幾次發行查詢未完成，已停止產出避免錯檔：\n" + "\n".join(missing))


def require_purposes(records: list[dict[str, str]], focus_keys: set[str] | None) -> None:
    missing: list[str] = []
    for record in records:
        if focus_keys is not None and record_key(record) not in focus_keys:
            continue
        if record.get("本次籌資計畫", "").strip():
            continue
        missing.append(
            f"{record.get('證券代號', '')} {record.get('公司名稱', '')} "
            f"{record.get('案件類別', '')} {record.get('金額', '')} "
            f"收文{record.get('收文日期', '')}"
        )
    if missing:
        raise ValueError("MOPS 本次籌資計畫原因查詢未完成，已停止產出避免錯檔：\n" + "\n".join(missing))


def complete_ordinals(ordinals: list[str], count: int) -> list[str]:
    cleaned: list[str] = []
    for ordinal in ordinals:
        value = chinese_ordinal_to_short(ordinal)
        if value and value not in cleaned:
            cleaned.append(value)
    if len(cleaned) >= count:
        return cleaned[:count]
    if not cleaned:
        return []
    numbers = [ordinal_to_int(value) for value in cleaned]
    numbers = [number for number in numbers if number is not None]
    start = min(numbers) if numbers else 1
    while len(cleaned) < count:
        candidate = int_to_chinese_ordinal(start + len(cleaned))
        if candidate not in cleaned:
            cleaned.append(candidate)
        else:
            start += 1
    return cleaned[:count]


def clean_explicit_bond_short(value: str) -> str:
    text = normalize_header(value)
    text = re.split(r"[)）(（，,。；;：:\\s]", text)[0]
    if not text or len(text) > 14:
        return ""
    bad_words = ("公告", "董事", "決議", "辦理", "發行", "公司", "債券", "代碼", "名稱", "資訊", "訊息")
    if any(word in text for word in bad_words):
        return ""
    if re.search(r"\d{6,}", text):
        return ""
    return text


def short_base_for_bond(record: dict[str, str], company_short: str = "") -> str:
    base = company_short or record.get("顯示名稱", "") or display_name_for_record(record) or record.get("公司名稱", "")
    return normalize_stock_short_for_bond(base)


def parse_bond_short_from_announcement(text: str, record: dict[str, str], company_short: str = "") -> str:
    matched_text = matching_record_text(text, record)
    compact = normalize_header(html_to_text(matched_text or text))
    if not compact:
        return ""
    base = short_base_for_bond(record, company_short)
    search_compact = compact
    anchors = [
        company_short,
        base,
        record.get("顯示名稱", ""),
        record.get("公司名稱", ""),
        record.get("證券代號", ""),
    ]
    anchor_positions = [
        search_compact.find(normalize_header(anchor))
        for anchor in anchors
        if normalize_header(anchor)
    ]
    anchor_positions = [pos for pos in anchor_positions if pos >= 0]
    if anchor_positions:
        pos = min(anchor_positions)
        search_compact = search_compact[pos : pos + 6000]
    for match in re.finditer(r"(?:債券簡稱|簡稱)[：:]\s*([^，,。；;\s()（）]{2,24})", compact):
        candidate = clean_explicit_bond_short(match.group(1))
        if candidate:
            return normalize_bond_product_name(candidate, record)
    bond_patterns = (
        r"(?:國內|海外)?(?:第)?([一二三四五六七八九十百\d]+)次[^。；，]{0,160}?(?:(?:有|無)?擔保)?(?:可)?(?:轉換|交換)公司債",
        r"(?:發行|辦理|募集)[^。；，]{0,160}?(?:國內|海外)?(?:第)?([一二三四五六七八九十百\d]+)次[^。；，]{0,160}?(?:轉換|交換)公司債",
        r"(?:轉換|交換)公司債[^。；，]{0,160}?(?:國內|海外)?(?:第)?([一二三四五六七八九十百\d]+)次",
        r"(?:國內|海外)?(?:第)?([一二三四五六七八九十百\d]+)次[^。；，]{0,160}?公司債",
    )
    for bond_pattern in bond_patterns:
        for match in re.finditer(bond_pattern, search_compact):
            ordinal = chinese_ordinal_to_short(match.group(1))
            if base:
                return bond_name_with_ordinal(base, ordinal, record)
    return ""


# 募資用途欄位名稱：CI 用「本次增資資金用途」，CB/ECB 用「募得價款之用途及運用計畫」
_PURPOSE_RE = re.compile(
    r"(?:本次增資資金用途|募得價款之用途及運用計畫|本次募集與發行資金用途|資金用途)[:：]\s*(.+?)"
    r"(?=。|\s*1[0-9]\.|\s*[0-9]+\.\s*其他|以上資料)",
    re.S,
)


def _purpose_fetch(path_qs: str) -> str:
    # MOPS 在高並行下會間歇 timeout，重試兩輪 + 較長 timeout 確保穩定抓到
    for attempt in range(3):
        for host in ("https://mopsov.twse.com.tw", "https://mops.twse.com.tw"):
            try:
                text = public_fetch_text(host + path_qs, timeout=6)
                if text and len(text) > 200:
                    return text
            except Exception:
                continue
    return ""


_RESOLUTION_CACHE: dict[tuple, list[tuple[str, str]]] = {}


def _company_resolution_announcements(record: dict[str, str], end: dt.date) -> list[tuple[str, str]]:
    """找該公司『董事會決議辦理現增/發行轉換公司債』公告，回傳 [(標題, 公告全文)]。

    這是名稱（第X次）與籌資計畫的單一可靠來源——以 co_id 精準鎖定該公司，
    避免主旨全文檢索抓到別家公司而張冠李戴。董事會決議日通常早於收文日 1-4 個月，
    故往前回溯 5 個月查找。
    """
    co_id = record.get("證券代號", "")
    classification = record.get("分類", "")
    received = parse_date(record.get("收文日期", "")) or end
    cache_key = (co_id, classification, received.year, received.month)
    cached = _RESOLUTION_CACHE.get(cache_key)
    if cached is not None:
        return cached

    typek = "sii" if "上市" in record.get("公司型態", "") else "otc"
    is_bond = classification in ("CB", "ECB", "EB")

    def title_matches(title: str) -> bool:
        # 排除非「董事會決議發行」的後續公告（收足價款、掛牌、賣回、終止等無資金用途段落）
        if any(bad in title for bad in ("減資", "限制員工", "子公司", "轉投資",
                                         "收足", "掛牌", "賣回", "上市買賣", "終止", "代收", "存儲", "價款")):
            return False
        if "董事會" not in title:
            return False
        if is_bond:
            return "轉換公司債" in title and ("發行" in title or "募集" in title)
        return "現金增資" in title and ("發行新股" in title or "發行普通股" in title)

    results: list[tuple[str, str]] = []
    first = received.replace(day=1)
    for back in range(0, 5):
        month = first.month - back
        year = first.year
        while month <= 0:
            month += 12
            year -= 1
        params = {
            "encodeURIComponent": "1", "step": "1", "firstin": "true", "off": "1",
            "TYPEK": typek, "co_id": co_id,
            "year": f"{year - 1911:03d}", "month": f"{month:02d}",
            "b_date": "1", "e_date": "31",
        }
        raw = _purpose_fetch("/mops/web/ajax_t05st01?" + urllib.parse.urlencode(params))
        if not raw:
            continue
        for match in re.finditer(r'onclick="([^"]*seq_no\.value[^"]+)"', raw):
            # MOPS 公告主旨常含換行，去除空白避免「轉換 公司債」斷字比對失敗
            title = re.sub(r"\s+", "", html_to_text(raw[max(0, match.start() - 300):match.start()]))
            if not title_matches(title):
                continue
            onclick = match.group(1)

            def field(name: str) -> str:
                found = re.search(name + r"\.value='([^']*)'", onclick)
                return found.group(1) if found else ""

            detail_params = {
                "step": "2", "TYPEK": field("TYPEK") or typek, "co_id": co_id,
                "seq_no": field("seq_no"), "spoke_time": field("spoke_time"),
                "spoke_date": field("spoke_date"), "firstin": "true",
            }
            detail = _purpose_fetch("/mops/web/ajax_t05st01?" + urllib.parse.urlencode(detail_params))
            if detail:
                results.append((title, html_to_text(detail)))
        if results:
            break  # 最近月份找到即停，避免抓到更舊的同類公告

    # 只快取非空結果：並行 timeout 的空結果不可快取，否則 sequential 補查會命中空 cache 而不重查
    if results:
        _RESOLUTION_CACHE[cache_key] = results
    return results


# 標題擷取本次發行序號，例如「發行國內第九次無擔保轉換公司債」-> 九
_ORDINAL_IN_TITLE_RE = re.compile(
    r"第\s*([一二三四五六七八九十百千\d]+)\s*次[^，。；]{0,12}(?:轉換公司債|交換公司債|公司債)"
)


def mops_bond_short_name_v2(record: dict[str, str], end: dt.date, company_short: str = "") -> str:
    """從該公司董事會決議發行公告『標題』擷取第X次，組合債券簡稱（如 世紀鋼九、台燿ECB5）。"""
    if record.get("分類") not in ("CB", "ECB", "EB"):
        return ""
    short = company_short or public_company_short_name(record.get("證券代號", "")) or record.get("公司名稱", "")
    for title, _detail in _company_resolution_announcements(record, end):
        m = _ORDINAL_IN_TITLE_RE.search(title)
        if m:
            ordinal = chinese_ordinal_to_short(m.group(1))
            name = bond_name_with_ordinal(short, ordinal, record)
            if name and is_bond_product_name(name):
                return name
    return ""


def mops_funding_purpose(record: dict[str, str], end: dt.date) -> str:
    """從董事會決議公告全文擷取資金用途。"""
    for _title, detail in _company_resolution_announcements(record, end):
        found = _PURPOSE_RE.search(detail)
        if found:
            return re.sub(r"\s+", "", found.group(1)).strip("、,，;；")
    # Fallback: open-data summary
    return funding_purpose_from_open_text(record, open_data_major_announcement_text(record))


def display_name_for_record(record: dict[str, str]) -> str:
    key = (record.get("證券代號", ""), record.get("分類", ""), record.get("金額", ""))
    if key in BOND_SHORT_NAMES:
        return BOND_SHORT_NAMES[key]
    security_code = record.get("證券代號", "")
    if security_code in SECURITY_SHORT_NAMES:
        return SECURITY_SHORT_NAMES[security_code]
    company = record.get("公司名稱", "").strip()
    if company.startswith("F-") and len(company) > 2:
        return f"{company[2:]}-KY"
    return company


def online_mops_lookup(record: dict[str, str], end: dt.date, need_company_short: bool, need_bond_name: bool, need_purpose: bool) -> dict[str, str]:
    result: dict[str, str] = {}
    short_name = ""
    if need_company_short:
        short_name = public_company_short_name(record.get("證券代號", ""))
        if short_name:
            result["company_short"] = short_name
            record["顯示名稱"] = short_name

    if need_bond_name:
        name = mops_bond_short_name_v2(record, end, short_name)
        if name:
            result["bond_name"] = name
            record["顯示名稱"] = name

    if need_purpose:
        purpose = mops_funding_purpose(record, end)
        if purpose:
            result["purpose"] = purpose
    return result


def enrich_records(
    records: list[dict[str, str]],
    end: dt.date | None = None,
    focus_keys: set[str] | None = None,
    purpose_keys: set[str] | None = None,
) -> list[str]:
    warnings: list[str] = []
    end = end or current_friday()
    lookup_jobs: list[tuple[dict[str, str], tuple[str, str, str], bool, bool, bool]] = []
    for record in records:
        key = (record.get("證券代號", ""), record.get("分類", ""), record.get("金額", ""))
        data = MOPS_ENRICHMENTS.get(key)
        is_focus = focus_keys is None or record_key(record) in focus_keys
        wants_fresh_purpose = purpose_keys is None or record_key(record) in purpose_keys
        if is_focus:
            record["顯示名稱"] = normalize_stock_short_for_bond(display_name_for_record(record))
            if wants_fresh_purpose:
                record["本次籌資計畫"] = ""
        elif not data:
            record["顯示名稱"] = normalize_stock_short_for_bond(display_name_for_record(record))
        else:
            record["顯示名稱"] = data["display"]
            record["本次籌資計畫"] = data["purpose"]
        if focus_keys is not None and record_key(record) not in focus_keys:
            continue
        # If cache already has a display name and we don't need fresh purpose, skip MOPS.
        if data and data.get("display") and not wants_fresh_purpose:
            record["顯示名稱"] = data["display"]
            record["本次籌資計畫"] = data.get("purpose", record.get("本次籌資計畫", ""))
            continue
        need_company_short = True
        need_bond_name = record.get("分類") in ("CB", "ECB", "EB")
        need_purpose = wants_fresh_purpose
        if not ENABLE_ONLINE_MOPS_LOOKUP:
            warnings.append(f"{record.get('證券代號')} {record.get('公司名稱')}：線上查詢未啟用；本工具要求線上查詢，請開啟 ENABLE_ONLINE_MOPS_LOOKUP。")
            if not need_bond_name:
                record["顯示名稱"] = normalize_stock_short_for_bond(display_name_for_record(record))
            if need_bond_name:
                record["顯示名稱"] = stock_display_name(record)
                warnings.append(f"{record.get('證券代號')} {record.get('公司名稱')}：未查詢 MOPS 第幾次名稱，請勿直接採用備援名稱。")
            if need_purpose:
                warnings.append(f"{record.get('證券代號')} {record.get('顯示名稱') or record.get('公司名稱')}：未查詢 MOPS 本次籌資計畫，已留白避免查錯。")
            continue
        lookup_jobs.append((record, key, need_company_short, need_bond_name, need_purpose))
    if not ENABLE_ONLINE_MOPS_LOOKUP:
        normalize_weekly_stock_names(records, focus_keys)
    if ENABLE_ONLINE_MOPS_LOOKUP and lookup_jobs:
        workers = max(1, min(MAX_MOPS_WORKERS, len(lookup_jobs)))
        with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(online_mops_lookup, record.copy(), end, need_company_short, need_bond_name, need_purpose): (record, key, need_company_short, need_bond_name, need_purpose)
                for record, key, need_company_short, need_bond_name, need_purpose in lookup_jobs
            }
            for future in concurrent.futures.as_completed(futures):
                record, key, need_company_short, need_bond_name, need_purpose = futures[future]
                try:
                    result = future.result()
                except Exception as exc:
                    warnings.append(f"{record.get('證券代號')} {record.get('顯示名稱') or record.get('公司名稱')}：MOPS 查詢失敗，請人工確認。{exc}")
                    continue
                if need_company_short:
                    if result.get("company_short"):
                        record["顯示名稱"] = result["company_short"]
                    else:
                        warnings.append(f"{record.get('證券代號')} {record.get('公司名稱')}：TWSE/TPEx 查不到公司簡稱，已先用來源檔公司名稱。")
                if need_bond_name:
                    if result.get("bond_name"):
                        record["顯示名稱"] = result["bond_name"]
                    else:
                        warnings.append(f"{record.get('證券代號')} {record.get('公司名稱')}：MOPS 查不到可確認的 CB/ECB 第幾次名稱。")
                if need_purpose:
                    if result.get("purpose"):
                        record["本次籌資計畫"] = result["purpose"]
                    else:
                        warnings.append(f"{record.get('證券代號')} {record.get('顯示名稱') or record.get('公司名稱')}：MOPS 找不到可精準比對的本次籌資計畫，已留白避免查錯，請人工確認。")
                # Persist successful CB/ECB lookups so next run skips MOPS for known records
                if key[1] in ("CB", "ECB", "EB") and record.get("顯示名稱"):
                    MOPS_ENRICHMENTS[key] = {
                        "display": record["顯示名稱"],
                        "purpose": record.get("本次籌資計畫", MOPS_ENRICHMENTS.get(key, {}).get("purpose", "")),
                    }
        resolve_missing_bond_names(records, end, focus_keys, warnings)
        # Sequential 補查：MOPS 在並行下少數筆會間歇 timeout 或 detail 截斷。
        # 對仍缺名稱/用途的筆，先清掉其快取強制重抓完整公告，無並行壓力下幾乎必成功。
        for record in records:
            rk = record_key(record)
            if focus_keys is not None and rk not in focus_keys:
                continue
            need_bond = record.get("分類") in ("CB", "ECB", "EB") and not is_bond_product_name(record.get("顯示名稱", ""))
            wants_purpose = purpose_keys is None or rk in purpose_keys
            need_purpose = wants_purpose and not record.get("本次籌資計畫", "").strip()
            if not need_bond and not need_purpose:
                continue
            received = parse_date(record.get("收文日期", "")) or end
            code = record.get("證券代號", "")
            cache_key = (record.get("證券代號", ""), record.get("分類", ""), received.year, received.month)
            # 線上 Render 固定 IP 易被 MOPS 短時限流。對仍缺的筆帶間隔重試，熬過限流窗口。
            for attempt in range(5):
                _RESOLUTION_CACHE.pop(cache_key, None)  # 清半殘快取強制重抓
                if need_bond:
                    name = mops_bond_short_name_v2(record, end)
                    if name:
                        record["顯示名稱"] = name
                        key = (record.get("證券代號", ""), record.get("分類", ""), record.get("金額", ""))
                        MOPS_ENRICHMENTS[key] = {"display": name, "purpose": record.get("本次籌資計畫", MOPS_ENRICHMENTS.get(key, {}).get("purpose", ""))}
                        warnings[:] = [w for w in warnings if not (w.split("：")[0].split()[0] == code and "第幾次名稱" in w)]
                        need_bond = False
                if need_purpose:
                    purpose = mops_funding_purpose(record, end)
                    if purpose:
                        record["本次籌資計畫"] = purpose
                        warnings[:] = [w for w in warnings if not (w.split("：")[0].split()[0] == code and "本次籌資計畫" in w)]
                        need_purpose = False
                if not need_bond and not need_purpose:
                    break
                _time.sleep(3)  # 等 MOPS 限流窗口恢復後再試
        reassign_multi_bond_ordinals(records, end, focus_keys, warnings)
        normalize_weekly_stock_names(records, focus_keys)
        require_bond_names(records, focus_keys)
        require_purposes(records, purpose_keys)
        _save_bond_cache()
    return warnings


def clean_broker(value: str) -> str:
    return re.sub(r"\s+", "", value).strip("、,，")


def normalize_broker(value: str) -> str:
    """正規化承銷商名供比對：源檔帶『證券』，樣板常用簡稱（永豐金證、第一金證）。
    去掉結尾的『證券』再去掉單一『證』或『券』，讓兩邊核心名對齊。"""
    name = clean_broker(value)
    name = re.sub(r"證券$", "", name)
    name = re.sub(r"[證券]$", "", name)
    return name


def in_range(value: str, start: dt.date, end: dt.date) -> bool:
    parsed = parse_date(value)
    return bool(parsed and start <= parsed <= end)


def unique_records(records: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[str] = set()
    unique: list[dict[str, str]] = []
    for record in records:
        key = record_key(record)
        if key in seen:
            continue
        seen.add(key)
        unique.append(record)
    return unique


def accepted_case(case_type: str) -> bool:
    text = normalize_header(case_type)
    return any(keyword in text for keyword in CASE_KEYWORDS)


def accepted_company_type(company_type: str) -> bool:
    return any(kind in normalize_header(company_type) for kind in COMPANY_TYPES)


def extract_records(xlsx_path: Path) -> list[dict[str, str]]:
    sheets = read_xlsx(xlsx_path)
    if not sheets:
        raise ValueError("來源 Excel 沒有工作表。")
    sheet = max(sheets, key=lambda item: len(item["rows"]))  # type: ignore[arg-type]
    rows = sheet["rows"]  # type: ignore[assignment]
    if is_imp_summary(rows):
        return extract_imp_summary_records(rows)
    header_idx, header = find_header(rows)
    purpose_cols = ["T", "U", "V", "W", "X", "Y", "Z", "AA", "AB"]
    records: list[dict[str, str]] = []
    for row in rows[header_idx + 1 :]:
        company = row_value(row, header, "公司名稱")
        case_type = row_value(row, header, "案件類別")
        company_type = row_value(row, header, "公司型態")
        if not company or not accepted_company_type(company_type) or not accepted_case(case_type):
            continue
        purpose = "；".join(v for col in purpose_cols if (v := row.get(col, "").strip()))
        close_type = row_value(row, header, "結案類型")
        amend_date = row.get("K", "").strip()
        stop_date = row.get("L", "").strip()
        release_date = row.get("M", "").strip()
        effective_date = row.get("N", "").strip()
        if close_type == "生效" and not effective_date:
            effective_date = amend_date
            amend_date = ""
        elif "停止生效" in close_type and not stop_date:
            stop_date = amend_date
            amend_date = ""

        record = {
            "證券代號": row_value(row, header, "證券代號"),
            "公司型態": company_type,
            "結案類型": close_type,
            "公司名稱": company,
            "承銷商": clean_broker(row_value(row, header, "承銷商")),
            "案件類別": case_type,
            "分類": case_code(case_type),
            "金額": row_value(row, header, "金額"),
            "幣別": row_value(row, header, "幣別"),
            "發行價格": row_value(row, header, "發行價格"),
            "收文日期": row_value(row, header, "收文日期"),
            "自動補正日期": row_value(row, header, "自動補正日期"),
            "停止生效日期": row_value(row, header, "停止生效日期"),
            "解除生效日期": row_value(row, header, "解除生效日期"),
            "生效日期": row_value(row, header, "生效日期"),
            "廢止/撤銷日期": row_value(row, header, "廢止撤銷日期"),
            "自行撤回日期": row_value(row, header, "自行撤回日期"),
            "退件日期": row_value(row, header, "退件日期"),
            "案件性質": row_value(row, header, "案件性質"),
            "承銷方式": row_value(row, header, "承銷方式"),
            "本次籌資計畫": purpose,
        }
        records.append(record)
    return records


def is_imp_summary(rows: list[dict[str, str]]) -> bool:
    if not rows:
        return False
    title = rows[0].get("A", "")
    header = next((row for row in rows[:6] if normalize_header(row.get("A", "")) == "證券代號"), {})
    return "申報案件辦理情形彙總表" in title and normalize_header(header.get("B", "")) == "公司型態"


def extract_imp_summary_records(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    status_words = ("生效", "自行撤回", "停止生效", "解除生效", "廢止", "撤銷", "退件")
    for row in rows[3:]:
        security_code = row.get("A", "").strip()
        company_type = row.get("B", "").strip()
        if not security_code or not accepted_company_type(company_type):
            continue
        c_value = row.get("C", "").strip()
        d_value = row.get("D", "").strip()
        close_type = c_value if d_value or c_value in status_words else ""
        company = d_value or ("" if c_value in status_words else c_value)
        broker = clean_broker(row.get("E", ""))
        case_type = row.get("F", "").strip()
        if not company or not broker or not accepted_case(case_type):
            continue

        code = case_code(case_type)
        currency = row.get("H", "").strip()
        issue_price = row.get("I", "").strip()
        received_date = row.get("J", "").strip()

        amend_date = row.get("K", "").strip()
        stop_date = row.get("L", "").strip()
        release_date = row.get("M", "").strip()
        effective_date = row.get("N", "").strip()
        if close_type == "生效" and not effective_date:
            effective_date = amend_date
            amend_date = ""
        elif "停止生效" in close_type and not stop_date:
            stop_date = amend_date
            amend_date = ""

        record = {
            "證券代號": security_code,
            "公司型態": company_type,
            "結案類型": close_type,
            "公司名稱": company,
            "承銷商": broker,
            "案件類別": case_type,
            "分類": code,
            "金額": row.get("G", "").strip(),
            "幣別": currency,
            "發行價格": issue_price,
            "收文日期": received_date,
            "自動補正日期": amend_date,
            "停止生效日期": stop_date,
            "解除生效日期": release_date,
            "生效日期": effective_date,
            "廢止/撤銷日期": "",
            "自行撤回日期": "",
            "退件日期": "",
            "案件性質": row.get("O", "").strip() or row.get("R", "").strip(),
            "承銷方式": "",
            "本次籌資計畫": "",
        }
        records.append(record)
    return records


def fetch_text(url: str) -> str:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(req, timeout=12) as resp:
            raw = resp.read()
            charset = resp.headers.get_content_charset() or "utf-8"
        return raw.decode(charset, errors="replace")
    except Exception:
        raw = fetch_with_windows(url)
        for charset in ("utf-8-sig", "utf-8", "big5", "cp950"):
            try:
                return raw.decode(charset)
            except UnicodeDecodeError:
                continue
        return raw.decode("utf-8", errors="replace")


def fetch_bytes(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(req, timeout=25) as resp:
            return resp.read()
    except Exception:
        return fetch_with_windows(url)


def fetch_with_windows(url: str, attempt_timeout: int = 25) -> bytes:
    errors: list[str] = []
    with tempfile.NamedTemporaryFile(delete=False) as tmp:
        tmp_path = tmp.name
    try:
        if attempt_timeout <= 8:
            methods = [
                ("curl.exe", ["curl.exe", "-L", "--retry", "0", "--connect-timeout", str(max(2, attempt_timeout // 2)), "--max-time", str(attempt_timeout), "-A", "Mozilla/5.0", "-o", tmp_path, url]),
            ]
        else:
            methods = [
                ("PowerShell Invoke-WebRequest", powershell_download_command(url, tmp_path, "iwr")),
                ("PowerShell WebClient", powershell_download_command(url, tmp_path, "webclient")),
                ("curl.exe", ["curl.exe", "-L", "--retry", "1", "--connect-timeout", "8", "--max-time", str(attempt_timeout), "-A", "Mozilla/5.0", "-o", tmp_path, url]),
                ("certutil", ["certutil.exe", "-urlcache", "-split", "-f", url, tmp_path]),
            ]
        for label, command in methods:
            try:
                completed = subprocess.run(command, check=True, timeout=attempt_timeout, capture_output=True)
                data = Path(tmp_path).read_bytes()
                if data:
                    return data
                errors.append(f"{label}: downloaded empty file")
            except Exception as exc:
                message = str(exc)
                stderr = getattr(exc, "stderr", b"")
                if stderr:
                    try:
                        message += " " + stderr.decode("utf-8", errors="ignore").strip()
                    except Exception:
                        pass
                errors.append(f"{label}: {message}")
        raise RuntimeError("自動下載失敗；已嘗試 Python、PowerShell、curl、certutil。請確認這台電腦允許連到證期局網站，或先用下方手動上傳來源檔。")
    finally:
        try:
            Path(tmp_path).unlink()
        except FileNotFoundError:
            pass


def powershell_download_command(url: str, target: str, mode: str) -> list[str]:
    prefix = (
        "[Net.ServicePointManager]::SecurityProtocol = "
        "[Net.SecurityProtocolType]::Tls12 -bor [Net.SecurityProtocolType]::Tls13; "
        "$ProgressPreference='SilentlyContinue'; "
    )
    if mode == "webclient":
        script = prefix + (
            "$wc = New-Object Net.WebClient; "
            "$wc.Headers.Add('User-Agent','Mozilla/5.0'); "
            f"$wc.DownloadFile({ps_quote(url)}, {ps_quote(target)})"
        )
    else:
        script = prefix + (
            f"Invoke-WebRequest -UseBasicParsing -Headers @{{'User-Agent'='Mozilla/5.0'}} "
            f"-Uri {ps_quote(url)} -OutFile {ps_quote(target)}"
        )
    return ["powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script]


def ps_quote(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def run_download_diagnostics() -> list[dict[str, str]]:
    results: list[dict[str, str]] = []
    for name, url in (("證期局頁面", SFB_PAGE), ("金管會網域", "https://www.fsc.gov.tw/")):
        try:
            data = fetch_with_windows(url)
            results.append({"name": name, "status": "成功", "detail": f"已下載 {len(data)} bytes"})
        except Exception as exc:
            results.append({"name": name, "status": "失敗", "detail": str(exc)})
    return results


def find_source_url(page_html: str, year: int) -> str:
    decoded = html.unescape(page_html)
    links = re.findall(r'<a\b[^>]*href="([^"]+)"[^>]*>(.*?)</a>', decoded, flags=re.I | re.S)
    candidates: list[tuple[int, str]] = []
    for href, label in links:
        text = re.sub(r"<[^>]+>", "", label)
        nearby_start = max(decoded.find(href) - 180, 0)
        nearby = decoded[nearby_start : decoded.find(href) + len(href) + 80]
        haystack = html.unescape(text + " " + nearby)
        if "EXCEL" not in haystack.upper() and not re.search(r"\.xlsx?$", href, re.I):
            continue
        score = 0
        if f"{year}年度申報案件" in haystack:
            score += 100
        if "申報案件" in haystack:
            score += 20
        if str(year) in haystack:
            score += 10
        if re.search(r"\.xlsx?$", href, re.I):
            score += 5
        if score:
            candidates.append((score, urllib.parse.urljoin(SFB_PAGE, href)))
    if not candidates:
        raise ValueError("在證期局頁面找不到年度申報案件 Excel 連結。")
    candidates.sort(reverse=True)
    return candidates[0][1]


def download_latest_source(end: dt.date) -> tuple[Path, str]:
    page = fetch_text(SFB_PAGE)
    source_url = find_source_url(page, roc_year(end))
    filename = urllib.parse.unquote(Path(urllib.parse.urlparse(source_url).path).name) or "source.xlsx"
    target = SOURCE_DIR / filename
    target.write_bytes(fetch_bytes(source_url))
    return target, source_url


def xml_escape(text: object) -> str:
    return html.escape(str(text if text is not None else ""), quote=True)


def make_sheet_xml(rows: list[list[object]]) -> str:
    col_count = max((len(row) for row in rows), default=1)
    row_xml = []
    for r_idx, row in enumerate(rows, start=1):
        cells = []
        for c_idx in range(1, col_count + 1):
            value = row[c_idx - 1] if c_idx <= len(row) else ""
            ref = f"{number_to_col(c_idx)}{r_idx}"
            cells.append(
                f'<c r="{ref}" t="inlineStr"><is><t>{xml_escape(value)}</t></is></c>'
            )
        row_xml.append(f'<row r="{r_idx}">{"".join(cells)}</row>')
    return (
        f'<worksheet xmlns="{NS_MAIN}" xmlns:r="{NS_REL}">'
        f'<dimension ref="A1:{number_to_col(col_count)}{max(len(rows), 1)}"/>'
        "<sheetViews><sheetView workbookViewId=\"0\"/></sheetViews>"
        "<sheetFormatPr defaultRowHeight=\"18\"/>"
        f'<sheetData>{"".join(row_xml)}</sheetData>'
        "</worksheet>"
    )


def write_xlsx(path: Path, sheets: list[tuple[str, list[list[object]]]]) -> None:
    content_types = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">',
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>',
        '<Default Extension="xml" ContentType="application/xml"/>',
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>',
    ]
    for i in range(1, len(sheets) + 1):
        content_types.append(
            f'<Override PartName="/xl/worksheets/sheet{i}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        )
    content_types.append("</Types>")

    workbook_sheets = []
    workbook_rels = []
    for i, (name, _) in enumerate(sheets, start=1):
        safe_name = xml_escape(name[:31])
        workbook_sheets.append(f'<sheet name="{safe_name}" sheetId="{i}" r:id="rId{i}"/>')
        workbook_rels.append(
            f'<Relationship Id="rId{i}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            f'Target="worksheets/sheet{i}.xml"/>'
        )

    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<workbook xmlns="{NS_MAIN}" xmlns:r="{NS_REL}"><sheets>'
        + "".join(workbook_sheets)
        + "</sheets></workbook>"
    )
    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="{NS_PKG_REL}">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/></Relationships>'
    )
    workbook_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="{NS_PKG_REL}">'
        + "".join(workbook_rels)
        + "</Relationships>"
    )

    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", "".join(content_types))
        zf.writestr("_rels/.rels", root_rels)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)
        for i, (_, rows) in enumerate(sheets, start=1):
            zf.writestr(f"xl/worksheets/sheet{i}.xml", make_sheet_xml(rows))


OUTPUT_COLUMNS = [
    "證券代號",
    "公司型態",
    "結案類型",
    "公司名稱",
    "承銷商",
    "案件類別",
    "分類",
    "金額",
    "幣別",
    "發行價格",
    "收文日期",
    "自動補正日期",
    "停止生效日期",
    "解除生效日期",
    "生效日期",
    "廢止/撤銷日期",
    "自行撤回日期",
    "退件日期",
    "案件性質",
    "承銷方式",
    "本次籌資計畫",
]


def records_to_rows(records: list[dict[str, str]]) -> list[list[object]]:
    return [OUTPUT_COLUMNS] + [[record.get(col, "") for col in OUTPUT_COLUMNS] for record in records]


def weekly_update_rows(sections: list[tuple[str, str, list[dict[str, str]]]]) -> list[list[object]]:
    rows: list[list[object]] = [["更新項目", "依據日期欄位", *OUTPUT_COLUMNS]]
    for title, date_column, records in sections:
        for record in records:
            rows.append([title, date_column, *[record.get(col, "") for col in OUTPUT_COLUMNS]])
    if len(rows) == 1:
        rows.append(["本週無更新", "", *["" for _ in OUTPUT_COLUMNS]])
    return rows


def purpose_check_rows(records: list[dict[str, str]]) -> list[list[object]]:
    rows: list[list[object]] = [["公司名稱", "承銷商", "案件類別", "分類", "本次籌資計畫", "檢核結果"]]
    if not records:
        rows.append(["本週無新增案件", "", "", "", "", "不用補公開觀測站原因"])
        return rows
    for record in records:
        purpose = record.get("本次籌資計畫", "").strip()
        rows.append(
            [
                record.get("公司名稱", ""),
                record.get("承銷商", ""),
                record.get("案件類別", ""),
                record.get("分類", ""),
                purpose,
                "已有資料" if purpose else "需至公開觀測站補原因",
            ]
        )
    return rows


def workflow_check_rows(start: dt.date, end: dt.date, source_path: Path, source_url: str, counts: dict[str, int]) -> list[list[object]]:
    return [
        ["Word步驟", "自動化處理", "結果"],
        ["Step1 複製上週檔案並更改截至日期", "產出新檔名與摘要週期", f"{roc_date(start)}～{roc_date(end)}"],
        ["Step2 至證期局網站下載檔案", "優先自動下載；失敗可上傳來源檔", source_url or source_path.name],
        ["Step3 篩選公司型態與案件類別", "上市/上櫃；CI/CB/ECB/GDR/EB", f"篩選後 {counts['all']} 筆"],
        ["Step4 更新新增/補正/停止/生效", "依各日期欄位切本週區間", f"新增 {counts['new']}、補正 {counts['amend']}、停止 {counts['stop']}、生效 {counts['effective']}"],
        ["Step5 查詢新增案件籌資原因", "來源檔若有用途欄位會帶入；空白列列在籌資目的檢核", f"待補 {counts['missingPurpose']} 筆"],
        ["Step6 統計券商送件案件家數", "依承銷商與案件分類統計", "已產出承銷商統計頁"],
    ]


DETAIL_COLS = {
    "證券代號": "A",
    "公司型態": "B",
    "結案類型": "C",
    "公司名稱": "D",
    "承銷商": "E",
    "案件類別": "F",
    "金額": "G",
    "幣別": "H",
    "發行價格": "I",
    "收文日期": "J",
    "自動補正日期": "K",
    "停止生效日期": "L",
    "解除生效日期": "M",
    "生效日期": "N",
    "廢止/撤銷日期": "O",
    "自行撤回日期": "P",
    "退件日期": "Q",
    "案件性質": "R",
    "承銷方式": "S",
}


PURPOSE_COLS = ["T", "U", "V", "W", "X", "Y", "Z"]
BLUE_RGB = "FF0000FF"
BLACK_RGB = "FF000000"


def is_blue_rgb(value: str) -> bool:
    color = value.strip().upper()
    return color.endswith("0000FF") or color.endswith("0563C1")


def escape_xml_text(value: object) -> str:
    text = str(value if value is not None else "")
    text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)
    return html.escape(text, quote=True)


def inline_cell_xml(ref: str, value: object, style: str | None = None) -> str:
    style_part = f' s="{style}"' if style is not None else ""
    text = str(value if value is not None else "")
    space = ' xml:space="preserve"' if text.startswith(" ") or text.endswith(" ") or "\n" in text else ""
    return f'<c r="{ref}"{style_part} t="inlineStr"><is><t{space}>{escape_xml_text(text)}</t></is></c>'


CELL_XML_RE = re.compile(r"<c\b[^>]*/>|<c\b[^>]*>[\s\S]*?</c>")


def cell_ref_from_xml(cell_xml: str) -> str:
    match = re.search(r'\br="([^"]+)"', cell_xml)
    return match.group(1) if match else ""


def upsert_cell_xml(row_xml: str, row_num: int, col: str, value: object, style: str | None = None) -> str:
    ref = f"{col}{row_num}"
    new_cell = inline_cell_xml(ref, value, style)
    for match in CELL_XML_RE.finditer(row_xml):
        if cell_ref_from_xml(match.group(0)) == ref:
            return row_xml[: match.start()] + new_cell + row_xml[match.end() :]
    row_close = "</row>"
    insert_at = row_xml.rfind(row_close)
    for match in CELL_XML_RE.finditer(row_xml):
        current_col = cell_col(cell_ref_from_xml(match.group(0)))
        if current_col and col_to_number(current_col) > col_to_number(col):
            insert_at = match.start()
            break
    return row_xml[:insert_at] + new_cell + row_xml[insert_at:]


def get_row_xml(sheet_xml: str, row_num: int) -> str:
    match = re.search(rf'<row\b[^>]*\br="{row_num}"[^>]*>[\s\S]*?</row>', sheet_xml)
    return match.group(0) if match else ""


def replace_row_xml(sheet_xml: str, row_num: int, row_xml: str) -> str:
    return re.sub(rf'<row\b[^>]*\br="{row_num}"[^>]*>[\s\S]*?</row>', lambda _: row_xml, sheet_xml, count=1)


def replace_exact_row_xml(sheet_xml: str, old_row_xml: str, new_row_xml: str) -> str:
    return sheet_xml.replace(old_row_xml, new_row_xml, 1)


def append_row_xml(sheet_xml: str, row_xml: str) -> str:
    return sheet_xml.replace("</sheetData>", row_xml + "</sheetData>", 1)


def clone_row_xml(row_xml: str, old_num: int, new_num: int) -> str:
    cloned = re.sub(rf'\b{old_num}\b', str(new_num), row_xml, count=1)
    cloned = re.sub(
        r'(<c\b[^>]*\br=")([A-Z]+)' + str(old_num) + r'(")',
        lambda m: f'{m.group(1)}{m.group(2)}{new_num}{m.group(3)}',
        cloned,
    )

    def clear_cell(match: re.Match[str]) -> str:
        cell_xml = match.group(0)
        tag_end = cell_xml.find(">")
        if tag_end < 0:
            return cell_xml
        opening = cell_xml[: tag_end + 1]
        if opening.endswith("/>"):
            opening = opening[:-2] + ">"
        opening = re.sub(r'\s+t="[^"]+"', "", opening)
        return opening[:-1].rstrip() + "/>"

    return CELL_XML_RE.sub(clear_cell, cloned)


def update_dimension_xml(sheet_xml: str, last_col: str, last_row: int) -> str:
    return re.sub(r'<dimension\b[^>]*ref="([^":]+)(?::[^"]+)?"[^>]*/>', rf'<dimension ref="\1:{last_col}{last_row}"/>', sheet_xml, count=1)


def decode_shared_text(raw: str) -> str:
    return html.unescape(raw or "")


def cell_value_from_xml(cell_xml: str, shared: list[str]) -> str:
    t_match = re.search(r'\bt="([^"]+)"', cell_xml)
    v_match = re.search(r'<v>([\s\S]*?)</v>', cell_xml)
    if t_match and t_match.group(1) == "inlineStr":
        return decode_shared_text("".join(re.findall(r'<t[^>]*>([\s\S]*?)</t>', cell_xml))).strip()
    value = decode_shared_text(v_match.group(1)) if v_match else ""
    if t_match and t_match.group(1) == "s" and value:
        idx = int(float(value))
        return shared[idx] if 0 <= idx < len(shared) else value
    return value.strip()


def row_values_from_xml(row_xml: str, shared: list[str]) -> dict[str, str]:
    values: dict[str, str] = {}
    for cell_xml in iter_cell_xml(row_xml):
        ref_match = re.search(r'\br="([A-Z]+)\d+"', cell_xml)
        if not ref_match:
            continue
        values[ref_match.group(1)] = cell_value_from_xml(cell_xml, shared)
    return values


def iter_cell_xml(row_xml: str):
    pos = 0
    while True:
        start = row_xml.find("<c", pos)
        if start < 0:
            break
        tag_end = row_xml.find(">", start)
        if tag_end < 0:
            break
        if row_xml[tag_end - 1] == "/":
            yield row_xml[start : tag_end + 1]
            pos = tag_end + 1
            continue
        end = row_xml.find("</c>", tag_end)
        if end < 0:
            break
        yield row_xml[start : end + 4]
        pos = end + 4


def cell_style_from_row(row_xml: str, col: str, row_num: int) -> str | None:
    ref = f"{col}{row_num}"
    for cell_xml in iter_cell_xml(row_xml):
        if re.search(rf'\br="{re.escape(ref)}"', cell_xml):
            match = re.search(r'\bs="([^"]+)"', cell_xml)
            return match.group(1) if match else None
    return None


def xml_items(container_xml: str, tag: str) -> list[str]:
    return re.findall(rf"<{tag}\b[^>]*(?:/>|>[\s\S]*?</{tag}>)", container_xml)


def set_opening_attr(xml: str, tag: str, attr: str, value: str) -> str:
    def replace(match: re.Match[str]) -> str:
        head, attrs, slash = match.group(1), match.group(2), match.group(3)
        if re.search(rf'\b{re.escape(attr)}="[^"]*"', attrs):
            attrs = re.sub(rf'\b{re.escape(attr)}="[^"]*"', f'{attr}="{value}"', attrs, count=1)
        else:
            attrs = attrs.rstrip() + f' {attr}="{value}"'
        return f"{head}{attrs}{slash}>"

    return re.sub(rf"^(<{tag}\b)([^>]*?)(/?)>", replace, xml, count=1)


def set_font_color(font_xml: str, rgb: str) -> str:
    color = f'<color rgb="{rgb}"/>'
    if re.search(r"<color\b[^>]*/>", font_xml):
        return re.sub(r"<color\b[^>]*/>", color, font_xml, count=1)
    if re.search(r"<color\b[^>]*></color>", font_xml):
        return re.sub(r"<color\b[^>]*></color>", color, font_xml, count=1)
    if re.search(r"<sz\b[^>]*/>", font_xml):
        return re.sub(r"(<sz\b[^>]*/>)", rf"\1{color}", font_xml, count=1)
    return font_xml.replace(">", ">" + color, 1)


def font_is_blue(font_xml: str) -> bool:
    for rgb in re.findall(r'<color\b[^>]*\brgb="([^"]+)"', font_xml, flags=re.I):
        if is_blue_rgb(rgb):
            return True
    return bool(re.search(r'<color\b[^>]*\bindexed="4"', font_xml, flags=re.I))


class XlsxStyleManager:
    def __init__(self, styles_xml: str):
        self.styles_xml = styles_xml
        fonts_match = re.search(r"<fonts\b[^>]*>[\s\S]*?</fonts>", styles_xml)
        cell_xfs_match = re.search(r"<cellXfs\b[^>]*>[\s\S]*?</cellXfs>", styles_xml)
        self.fonts = xml_items(fonts_match.group(0), "font") if fonts_match else []
        self.cell_xfs = xml_items(cell_xfs_match.group(0), "xf") if cell_xfs_match else []
        self.font_cache: dict[tuple[int, str], int] = {}
        self.style_cache: dict[tuple[int, int], int] = {}
        self.black_style_cache: dict[str | None, str | None] = {}
        self.blue_style_cache: dict[str | None, str | None] = {}

    def _font_id_for_style(self, style: str | None) -> int:
        try:
            style_id = int(style or "0")
            xf = self.cell_xfs[style_id]
        except (ValueError, IndexError):
            return 0
        match = re.search(r'\bfontId="(\d+)"', xf)
        return int(match.group(1)) if match else 0

    def _style_with_font(self, style: str | None, font_id: int) -> str | None:
        try:
            style_id = int(style or "0")
            base = self.cell_xfs[style_id]
        except (ValueError, IndexError):
            return style
        key = (style_id, font_id)
        if key in self.style_cache:
            return str(self.style_cache[key])
        new_xf = set_opening_attr(base, "xf", "fontId", str(font_id))
        new_xf = set_opening_attr(new_xf, "xf", "applyFont", "1")
        for index, xf in enumerate(self.cell_xfs):
            if xf == new_xf:
                self.style_cache[key] = index
                return str(index)
        self.cell_xfs.append(new_xf)
        new_id = len(self.cell_xfs) - 1
        self.style_cache[key] = new_id
        return str(new_id)

    def _font_with_color(self, font_id: int, rgb: str) -> int:
        key = (font_id, rgb)
        if key in self.font_cache:
            return self.font_cache[key]
        base = self.fonts[font_id] if 0 <= font_id < len(self.fonts) else (self.fonts[0] if self.fonts else "<font/>")
        new_font = set_font_color(base, rgb)
        for index, font in enumerate(self.fonts):
            if font == new_font:
                self.font_cache[key] = index
                return index
        self.fonts.append(new_font)
        new_id = len(self.fonts) - 1
        self.font_cache[key] = new_id
        return new_id

    def black_style(self, style: str | None) -> str | None:
        if style in self.black_style_cache:
            return self.black_style_cache[style]
        font_id = self._font_id_for_style(style)
        if 0 <= font_id < len(self.fonts) and font_is_blue(self.fonts[font_id]):
            black_font = self._font_with_color(font_id, BLACK_RGB)
            result = self._style_with_font(style, black_font)
        else:
            result = style
        self.black_style_cache[style] = result
        return result

    def blue_style(self, style: str | None) -> str | None:
        if style in self.blue_style_cache:
            return self.blue_style_cache[style]
        base_style = self.black_style(style)
        base_font = self._font_id_for_style(base_style)
        blue_font = self._font_with_color(base_font, BLUE_RGB)
        result = self._style_with_font(base_style, blue_font)
        self.blue_style_cache[style] = result
        return result

    def finalize(self) -> str:
        def replace_section(xml: str, tag: str, items: list[str]) -> str:
            def replace(match: re.Match[str]) -> str:
                opening = re.sub(r'\bcount="\d+"', f'count="{len(items)}"', match.group(1), count=1)
                return opening + "".join(items) + match.group(3)

            return re.sub(rf"(<{tag}\b[^>]*>)([\s\S]*?)(</{tag}>)", replace, xml, count=1)

        self.styles_xml = replace_section(self.styles_xml, "fonts", self.fonts)
        self.styles_xml = replace_section(self.styles_xml, "cellXfs", self.cell_xfs)
        return self.styles_xml


def append_shared_string(shared_xml: str, text: str, blue_suffix: str = "") -> tuple[str, int]:
    sis = re.findall(r"<si>[\s\S]*?</si>", shared_xml)
    index = len(sis)
    if blue_suffix:
        prefix = text[: -len(blue_suffix)] if text.endswith(blue_suffix) else text.replace(blue_suffix, "")
        si = (
            "<si>"
            f"<r><rPr><sz val=\"12\"/><name val=\"微軟正黑體\"/><family val=\"2\"/><charset val=\"136\"/></rPr><t>{escape_xml_text(prefix)}</t></r>"
            f"<r><rPr><sz val=\"12\"/><color rgb=\"FF0000FF\"/><name val=\"微軟正黑體\"/><family val=\"2\"/><charset val=\"136\"/></rPr><t>{escape_xml_text(blue_suffix)}</t></r>"
            "</si>"
        )
    else:
        si = f"<si><t>{escape_xml_text(text)}</t></si>"
    shared_xml = shared_xml.replace("</sst>", si + "</sst>", 1)
    shared_xml = re.sub(
        r'(<sst\b[^>]*\bcount=")(\d+)(")',
        lambda m: f"{m.group(1)}{int(m.group(2)) + 1}{m.group(3)}",
        shared_xml,
        count=1,
    )
    shared_xml = re.sub(
        r'(<sst\b[^>]*\buniqueCount=")(\d+)(")',
        lambda m: f"{m.group(1)}{int(m.group(2)) + 1}{m.group(3)}",
        shared_xml,
        count=1,
    )
    return shared_xml, index


def clear_previous_blue_runs(shared_xml: str) -> str:
    # Previous weekly highlights in the template are rich-text runs. Convert them
    # back to plain shared strings before adding this week's blue text.
    def replace_si(match: re.Match[str]) -> str:
        si = match.group(0)
        colors = re.findall(r'<color\b[^>]*\brgb="([^"]+)"', si, flags=re.I)
        indexed_blue = re.search(r'<color\b[^>]*\bindexed="4"', si, flags=re.I)
        if not indexed_blue and not any(is_blue_rgb(color) for color in colors):
            return si
        text = "".join(html.unescape(t) for t in re.findall(r"<t[^>]*>([\s\S]*?)</t>", si))
        return f"<si><t>{escape_xml_text(text)}</t></si>"

    return re.sub(r"<si>[\s\S]*?</si>", replace_si, shared_xml)


def clear_blue_cell_styles(sheet_xml: str, styles: XlsxStyleManager) -> str:
    def replace(match: re.Match[str]) -> str:
        return f'{match.group(1)}{styles.black_style(match.group(2))}{match.group(3)}'

    return re.sub(r'(<c\b[^>]*\bs=")(\d+)(")', replace, sheet_xml)


def upsert_shared_cell_xml(row_xml: str, row_num: int, col: str, shared_index: int, style: str | None = None) -> str:
    ref = f"{col}{row_num}"
    style_part = f' s="{style}"' if style is not None else ""
    new_cell = f'<c r="{ref}"{style_part} t="s"><v>{shared_index}</v></c>'
    for match in CELL_XML_RE.finditer(row_xml):
        if cell_ref_from_xml(match.group(0)) == ref:
            return row_xml[: match.start()] + new_cell + row_xml[match.end() :]
    insert_at = row_xml.rfind("</row>")
    for match in CELL_XML_RE.finditer(row_xml):
        current_col = cell_col(cell_ref_from_xml(match.group(0)))
        if current_col and col_to_number(current_col) > col_to_number(col):
            insert_at = match.start()
            break
    return row_xml[:insert_at] + new_cell + row_xml[insert_at:]


def set_cell_cached_value_xml(row_xml: str, row_num: int, col: str, value: object) -> str:
    ref = f"{col}{row_num}"

    def replace_cell(match: re.Match[str]) -> str:
        cell = match.group(0)
        value_xml = f"<v>{escape_xml_text(value)}</v>"
        if re.search(r"<v>[\s\S]*?</v>", cell):
            return re.sub(r"<v>[\s\S]*?</v>", value_xml, cell, count=1)
        return cell[:-4] + value_xml + "</c>" if cell.endswith("</c>") else cell

    pattern = re.compile(rf'<c\b[^>]*\br="{re.escape(ref)}"[^>]*(?:>[\s\S]*?</c>|/>)')
    return pattern.sub(replace_cell, row_xml, count=1)


def record_key(record: dict[str, str]) -> str:
    amount_digits = re.sub(r"\D", "", str(record.get("金額", "") or ""))
    parts = [
        record.get("證券代號", ""),
        record.get("分類", "") or case_code(record.get("案件類別", "")),
        amount_digits or record.get("金額", ""),
        record.get("收文日期", ""),
    ]
    return "|".join(normalize_header(part) for part in parts)


def detail_row_to_record(row: dict[str, str]) -> dict[str, str]:
    return {
        "證券代號": row.get("A", ""),
        "公司型態": row.get("B", ""),
        "結案類型": row.get("C", ""),
        "公司名稱": row.get("D", ""),
        "承銷商": clean_broker(row.get("E", "")),
        "案件類別": row.get("F", ""),
        "分類": case_code(row.get("F", "")),
        "金額": row.get("G", ""),
        "幣別": row.get("H", ""),
        "發行價格": row.get("I", ""),
        "收文日期": row.get("J", ""),
        "自動補正日期": row.get("K", ""),
        "停止生效日期": row.get("L", ""),
        "解除生效日期": row.get("M", ""),
        "生效日期": row.get("N", ""),
        "廢止/撤銷日期": row.get("O", ""),
        "自行撤回日期": row.get("P", ""),
        "退件日期": row.get("Q", ""),
        "案件性質": row.get("R", ""),
        "承銷方式": row.get("S", ""),
        "本次籌資計畫": "；".join(row.get(col, "").strip() for col in PURPOSE_COLS if row.get(col, "").strip()),
    }


def sheet_path_map(zf: zipfile.ZipFile) -> dict[str, str]:
    ns = {"m": NS_MAIN, "r": NS_REL, "pr": NS_PKG_REL}
    workbook = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rel_map = {}
    for rel in rels.findall("pr:Relationship", ns):
        target = rel.attrib.get("Target", "")
        rel_map[rel.attrib["Id"]] = posixpath.normpath(("xl/" + target).replace("xl//", "xl/"))
    paths = {}
    for sheet in workbook.findall("m:sheets/m:sheet", ns):
        rid = sheet.attrib.get(f"{{{NS_REL}}}id", "")
        paths[sheet.attrib.get("name", "")] = rel_map.get(rid, "")
    return paths


def find_sheet_path(paths: dict[str, str], preferred_name: str, fallback_keywords: tuple[str, ...], default_path: str) -> str:
    if preferred_name in paths:
        return paths[preferred_name]
    for name, path in paths.items():
        if all(keyword in name for keyword in fallback_keywords):
            return path
    return default_path


def worksheet_rows(root: ET.Element) -> list[ET.Element]:
    return root.findall(f".//{{{NS_MAIN}}}sheetData/{{{NS_MAIN}}}row")


def worksheet_dimension(root: ET.Element, last_row: int) -> None:
    dim = root.find(f"{{{NS_MAIN}}}dimension")
    if dim is not None:
        ref = dim.attrib.get("ref", "A1:Z1")
        start = ref.split(":")[0]
        dim.attrib["ref"] = f"{start}:Z{last_row}"


def cell_ref_col(ref: str) -> str:
    return re.sub(r"\d+", "", ref)


def row_cell_map(row: ET.Element) -> dict[str, ET.Element]:
    cells: dict[str, ET.Element] = {}
    for cell in row.findall(f"{{{NS_MAIN}}}c"):
        col = cell_ref_col(cell.attrib.get("r", ""))
        if col:
            cells[col] = cell
    return cells


def set_inline_cell(row: ET.Element, col: str, row_num: int, value: object, style: str | None = None) -> None:
    cells = row_cell_map(row)
    cell = cells.get(col)
    if cell is None:
        cell = ET.Element(f"{{{NS_MAIN}}}c", {"r": f"{col}{row_num}"})
        inserted = False
        for idx, existing in enumerate(list(row)):
            if existing.tag.endswith("c") and col_to_number(cell_ref_col(existing.attrib.get("r", ""))) > col_to_number(col):
                row.insert(idx, cell)
                inserted = True
                break
        if not inserted:
            row.append(cell)
    cell.attrib["r"] = f"{col}{row_num}"
    if style is not None:
        cell.attrib["s"] = style
    cell.attrib["t"] = "inlineStr"
    for child in list(cell):
        cell.remove(child)
    is_elem = ET.SubElement(cell, f"{{{NS_MAIN}}}is")
    t_elem = ET.SubElement(is_elem, f"{{{NS_MAIN}}}t")
    text = str(value if value is not None else "")
    if text.startswith(" ") or text.endswith(" ") or "\n" in text:
        t_elem.attrib["{http://www.w3.org/XML/1998/namespace}space"] = "preserve"
    t_elem.text = text


def row_values(row: ET.Element, shared: list[str]) -> dict[str, str]:
    values: dict[str, str] = {}
    for cell in row.findall(f"{{{NS_MAIN}}}c"):
        col = cell_ref_col(cell.attrib.get("r", ""))
        text = ""
        if cell.attrib.get("t") == "inlineStr":
            inline = cell.find(f"{{{NS_MAIN}}}is")
            text = xml_text(inline) if inline is not None else ""
        else:
            v = cell.find(f"{{{NS_MAIN}}}v")
            text = v.text if v is not None and v.text is not None else ""
            if cell.attrib.get("t") == "s" and text:
                idx = int(float(text))
                text = shared[idx] if 0 <= idx < len(shared) else text
        if col:
            values[col] = text.strip()
    return values


def xlsx_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    return [xml_text(si) for si in root.findall(f"{{{NS_MAIN}}}si")]


def clone_row(last_row: ET.Element, new_num: int) -> ET.Element:
    cloned = deepcopy(last_row)
    cloned.attrib["r"] = str(new_num)
    for cell in cloned.findall(f"{{{NS_MAIN}}}c"):
        col = cell_ref_col(cell.attrib.get("r", ""))
        cell.attrib["r"] = f"{col}{new_num}"
        for child in list(cell):
            cell.remove(child)
        cell.attrib.pop("t", None)
    return cloned


def split_purpose(record: dict[str, str]) -> list[str]:
    purpose = record.get("本次籌資計畫", "").strip()
    if not purpose:
        return ["", "", "", "", "", "", "待補事件原因"]
    columns = ["", "", "", "", "", "", ""]
    other: list[str] = []
    chunks = [chunk.strip() for chunk in re.split(r"[；;、,，]", purpose) if chunk.strip()]
    for chunk in chunks:
        compact = normalize_header(chunk)
        if "償還銀行借款" in compact or "償還金融機構借款" in compact:
            columns[0] = columns[0] or "償還銀行借款"
        elif "充實營運資金" in compact or "充實營運週轉金" in compact or "營運資金" in compact:
            columns[1] = columns[1] or "充實營運資金"
        elif "擴建廠房" in compact or "興建廠房" in compact:
            columns[2] = columns[2] or chunk
        elif "購置機器設備" in compact or "購置設備" in compact:
            columns[3] = columns[3] or chunk
        elif "轉投資" in compact or "投資國內外" in compact:
            columns[4] = columns[4] or chunk
        elif "償還前次" in compact or "償還第" in compact:
            columns[5] = columns[5] or chunk
        else:
            other.append(chunk)
    if other:
        columns[6] = "；".join(other)
    return columns


TRIGGER_DATE_FIELDS = {
    "自動補正日期": "K",
    "停止生效日期": "L",
    "解除生效日期": "M",
    "生效日期": "N",
    "廢止/撤銷日期": "O",
    "自行撤回日期": "P",
    "退件日期": "Q",
}
OPENPYXL_DETAIL_COLUMNS = {field: index for index, field in enumerate(DETAIL_COLS, start=1)}
OPENPYXL_PURPOSE_COLUMNS = list(range(20, 27))
OPENPYXL_SUMMARY_COLUMNS = {"CI": 51, "CB": 52, "ECB": 53, "GDR": 54, "EB": 55}
COMPARE_EVENT_BLUE_COLUMNS = {
    "自動補正日期": "K",
    "停止生效日期": "L",
    "解除生效日期": "M",
    "生效日期": "C",
    "廢止/撤銷日期": "O",
    "自行撤回日期": "P",
    "退件日期": "Q",
}


def font_color_is_blue(color: object) -> bool:
    if color is None:
        return False
    color_type = getattr(color, "type", None)
    if color_type == "rgb":
        return is_blue_rgb(str(getattr(color, "rgb", "") or ""))
    if color_type == "indexed":
        return getattr(color, "indexed", None) == 4
    return False


def set_cell_font_color(cell: object, rgb: str) -> None:
    font = copy_style(cell.font)
    font.color = rgb
    cell.font = font


def blue_cell(cell: object) -> None:
    set_cell_font_color(cell, BLUE_RGB)


def black_cell(cell: object) -> None:
    if font_color_is_blue(cell.font.color):
        set_cell_font_color(cell, BLACK_RGB)


def clear_workbook_blue_fonts(workbook: object) -> None:
    for worksheet in workbook.worksheets:
        max_col = min(worksheet.max_column or 1, 80)
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row or 1, max_col=max_col):
            for cell in row:
                black_cell(cell)


def safe_load_workbook(path: Path):
    try:
        return load_workbook(path)
    except TypeError:
        sanitized = Path(tempfile.gettempdir()) / f"sanitized_{dt.datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{path.name}"
        sanitize_shared_strings(path, sanitized)
        return load_workbook(sanitized)


def sanitize_shared_strings(source: Path, target: Path) -> None:
    with zipfile.ZipFile(source, "r") as zin, zipfile.ZipFile(target, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            data = zin.read(name)
            if name == "xl/sharedStrings.xml":
                text = data.decode("utf-8")

                def flatten_si(match: re.Match[str]) -> str:
                    si = match.group(0)
                    if "<r>" not in si:
                        return si
                    value = "".join(html.unescape(t) for t in re.findall(r"<t[^>]*>([\s\S]*?)</t>", si))
                    return f"<si><t>{escape_xml_text(value)}</t></si>"

                data = re.sub(r"<si>[\s\S]*?</si>", flatten_si, text).encode("utf-8")
            zout.writestr(name, data)


def copy_cell_style(source_cell: object, target_cell: object) -> None:
    target_cell.font = copy_style(source_cell.font)
    target_cell.fill = copy_style(source_cell.fill)
    target_cell.border = copy_style(source_cell.border)
    target_cell.alignment = copy_style(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy_style(source_cell.protection)


def find_openpyxl_detail_sheet(workbook: object, end: dt.date):
    preferred = f"{roc_year(end)}年本次籌資計畫"
    if preferred in workbook.sheetnames:
        return workbook[preferred]
    for sheet in workbook.worksheets:
        if "本次籌資計畫" in sheet.title:
            return sheet
    return workbook.worksheets[1] if len(workbook.worksheets) > 1 else workbook.active


def find_openpyxl_summary_sheet(workbook: object, end: dt.date):
    preferred = f"{roc_year(end)}年"
    if preferred in workbook.sheetnames:
        return workbook[preferred]
    return workbook.worksheets[0]


def detail_row_record_from_sheet(sheet: object, row_num: int) -> dict[str, str]:
    values = {
        col: "" if sheet[f"{col}{row_num}"].value is None else str(sheet[f"{col}{row_num}"].value).strip()
        for col in [number_to_col(i) for i in range(1, 27)]
    }
    return detail_row_to_record(values)


def last_detail_row(sheet: object) -> int:
    last = 4
    for row_num in range(4, sheet.max_row + 1):
        code = sheet.cell(row_num, 1).value
        name = sheet.cell(row_num, 4).value
        if str(code or "").strip() or str(name or "").strip():
            last = row_num
    return last


def write_detail_row(sheet: object, row_num: int, record: dict[str, str], blue_cols: set[int], template_row: int) -> None:
    for col_idx in range(1, 27):
        copy_cell_style(sheet.cell(template_row, col_idx), sheet.cell(row_num, col_idx))
        black_cell(sheet.cell(row_num, col_idx))
    for field, col_letter in DETAIL_COLS.items():
        cell = sheet[f"{col_letter}{row_num}"]
        cell.value = detail_field_value(record, field)
        if col_to_number(col_letter) in blue_cols and str(cell.value or "").strip():
            blue_cell(cell)
    purpose_values = split_purpose(record)
    for idx, col_idx in enumerate(OPENPYXL_PURPOSE_COLUMNS):
        cell = sheet.cell(row_num, col_idx)
        cell.value = purpose_values[idx] if idx < len(purpose_values) else ""
        if col_idx in blue_cols and str(cell.value or "").strip():
            blue_cell(cell)


def update_existing_detail_row(sheet: object, row_num: int, record: dict[str, str], blue_cols: set[int]) -> None:
    for field, col_letter in DETAIL_COLS.items():
        cell = sheet[f"{col_letter}{row_num}"]
        cell.value = detail_field_value(record, field)
        black_cell(cell)
        if col_to_number(col_letter) in blue_cols and str(cell.value or "").strip():
            blue_cell(cell)
    purpose_values = split_purpose(record)
    for idx, col_idx in enumerate(OPENPYXL_PURPOSE_COLUMNS):
        cell = sheet.cell(row_num, col_idx)
        cell.value = purpose_values[idx] if idx < len(purpose_values) else ""
        black_cell(cell)
        if col_idx in blue_cols and str(cell.value or "").strip():
            blue_cell(cell)


def changed_value(old: str, new: str) -> bool:
    old_norm = normalize_header(old)
    new_norm = normalize_header(new)
    return bool(new_norm) and old_norm != new_norm


def compare_blue_columns(previous: dict[str, str] | None, current: dict[str, str]) -> set[int]:
    if previous is None:
        return {col_to_number("D")}
    cols: set[int] = set()
    for field, col_letter in COMPARE_EVENT_BLUE_COLUMNS.items():
        if changed_value(previous.get(field, ""), current.get(field, "")):
            cols.add(col_to_number(col_letter))
    return cols


def weekly_blue_columns(record: dict[str, str], start: dt.date, end: dt.date) -> set[int]:
    cols: set[int] = set()
    if in_range(record.get("收文日期", ""), start, end):
        cols.add(col_to_number("D"))
    if in_range(record.get("自動補正日期", ""), start, end):
        cols.add(col_to_number("K"))
    if in_range(record.get("停止生效日期", ""), start, end):
        cols.add(col_to_number("L"))
    if in_range(record.get("解除生效日期", ""), start, end):
        cols.add(col_to_number("M"))
    if in_range(record.get("生效日期", ""), start, end):
        cols.add(col_to_number("C"))
    if in_range(record.get("廢止/撤銷日期", ""), start, end):
        cols.add(col_to_number("O"))
    if in_range(record.get("自行撤回日期", ""), start, end):
        cols.add(col_to_number("P"))
    if in_range(record.get("退件日期", ""), start, end):
        cols.add(col_to_number("Q"))
    return cols


def detail_field_value(record: dict[str, str], field: str) -> str:
    if field == "公司名稱":
        display = record.get("顯示名稱", "")
        if record.get("分類") in ("CB", "ECB", "EB") and is_bond_product_name(display):
            return display
        if record.get("分類") not in ("CB", "ECB", "EB") and display:
            return display
    return record.get(field, "")


def update_summary_sheet_openpyxl(
    sheet: object,
    base_records: list[dict[str, str]],
    new_records: list[dict[str, str]],
    start: dt.date,
    end: dt.date,
    weekly_keys: set[str] | None = None,
) -> None:
    sheet["BB1"] = f"更新日期：{roc_date(end)}"
    sheet["AX2"] = f"{roc_year(end)}.01.01~{roc_date(end)}"
    broker_rows: dict[str, int] = {}
    broker_orig: dict[str, str] = {}  # normalized -> 樣板原始承銷商名（供排序後寫回 A 欄）
    for row_num in range(4, sheet.max_row + 1):
        raw_name = str(sheet.cell(row_num, 1).value or "")
        broker = normalize_broker(raw_name)
        if broker:
            broker_rows[broker] = row_num
            broker_orig[broker] = clean_broker(raw_name)

    year_records = [r for r in base_records if (parse_date(r.get("收文日期", "")) or start) <= end]
    by_broker: dict[str, dict[str, list[str]]] = {}
    weekly_by_broker: dict[str, dict[str, list[str]]] = {}
    for record in year_records:
        broker = normalize_broker(record.get("承銷商", "")) or "未填"
        code = record.get("分類", "其他")
        if code not in OPENPYXL_SUMMARY_COLUMNS:
            continue
        by_broker.setdefault(broker, {key: [] for key in OPENPYXL_SUMMARY_COLUMNS})
        name = record.get("顯示名稱") or record.get("公司名稱", "")
        by_broker[broker][code].append(name)
    for record in year_records:
        if weekly_keys is not None:
            if record_key(record) not in weekly_keys:
                continue
        elif not weekly_blue_columns(record, start, end):
            continue
        broker = normalize_broker(record.get("承銷商", "")) or "未填"
        code = record.get("分類", "其他")
        if code not in OPENPYXL_SUMMARY_COLUMNS:
            continue
        weekly_by_broker.setdefault(broker, {key: [] for key in OPENPYXL_SUMMARY_COLUMNS})
        name = record.get("顯示名稱") or record.get("公司名稱", "")
        weekly_by_broker[broker][code].append(name)

    # 有送件的承銷商按件數由高到低排序，填入它們原本佔用的列（列位置/格式不變，內容降序對應）
    data_brokers = [
        (broker, code_map) for broker, code_map in by_broker.items()
        if broker in broker_rows and broker != "合計"
    ]
    data_brokers.sort(key=lambda item: sum(len(v) for v in item[1].values()), reverse=True)
    target_rows = sorted(broker_rows[broker] for broker, _ in data_brokers)
    for (broker, code_map), row_num in zip(data_brokers, target_rows):
        sheet.cell(row_num, 1).value = broker_orig.get(broker, broker)
        total = sum(len(items) for items in code_map.values())
        sheet.cell(row_num, 50).value = total
        for code, col_idx in OPENPYXL_SUMMARY_COLUMNS.items():
            names = [name for name in code_map.get(code, []) if name]
            weekly_names = set(weekly_by_broker.get(broker, {}).get(code, []))
            cell = sheet.cell(row_num, col_idx)
            cell.value = summary_rich_text(names, weekly_names)
            black_cell(cell)

    if "合計" in broker_rows:
        row_num = broker_rows["合計"]
        totals = {
            code: sum(len(code_map.get(code, [])) for code_map in by_broker.values())
            for code in OPENPYXL_SUMMARY_COLUMNS
        }
        sheet.cell(row_num, 50).value = sum(totals.values())
        for code, col_idx in OPENPYXL_SUMMARY_COLUMNS.items():
            sheet.cell(row_num, col_idx).value = totals[code]


def summary_rich_text(names: list[str], weekly_names: set[str]) -> object:
    if not names:
        return ""
    if not weekly_names:
        return "、".join(names)
    parts: list[object] = []
    blue_font = InlineFont(color=BLUE_RGB)
    for index, name in enumerate(names):
        prefix = "、" if index else ""
        text = prefix + name
        if name in weekly_names:
            parts.append(TextBlock(blue_font, text))
        else:
            parts.append(text)
    return CellRichText(*parts)


def update_template_workbook_openpyxl(
    source_path: Path,
    start: dt.date,
    end: dt.date,
    source_url: str,
    template_path: Path | None,
    base_records: list[dict[str, str]],
    counts: dict[str, int],
) -> Path:
    template = template_path if template_path and template_path.exists() else (TEMPLATE_PATH if TEMPLATE_PATH.exists() else source_path)
    filename = f"同業送件明細(截至{compact_roc_date(end)}).xlsx"
    target = available_report_path(REPORT_DIR / filename)
    workbook = safe_load_workbook(template)
    clear_workbook_blue_fonts(workbook)
    detail_sheet = find_openpyxl_detail_sheet(workbook, end)
    summary_sheet = find_openpyxl_summary_sheet(workbook, end)

    template_row = 4
    weekly_keys: set[str] = set()

    for idx, record in enumerate(base_records):
        row_num = 4 + idx
        blue_cols = weekly_blue_columns(record, start, end)
        if blue_cols:
            weekly_keys.add(record_key(record))
        write_detail_row(detail_sheet, row_num, record, blue_cols, template_row)

    # 清除樣板多餘的舊資料列
    for row_num in range(4 + len(base_records), (detail_sheet.max_row or 3) + 1):
        for col_idx in range(1, 27):
            detail_sheet.cell(row_num, col_idx).value = None

    new_records = [r for r in base_records if in_range(r.get("收文日期", ""), start, end)]
    update_summary_sheet_openpyxl(summary_sheet, base_records, new_records, start, end, weekly_keys)
    workbook.save(target)
    return target


def available_report_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for index in range(1, 100):
        candidate = path.with_name(f"{stem}-{index}{suffix}")
        if not candidate.exists():
            return candidate
    return path.with_name(f"{stem}-{dt.datetime.now().strftime('%H%M%S')}{suffix}")


def group_for_email(records: list[dict[str, str]]) -> list[str]:
    grouped: dict[str, list[str]] = {}
    for record in records:
        company = record.get("顯示名稱") or record.get("公司名稱", "")
        grouped.setdefault(record.get("分類", "其他"), []).append(
            f"{company}：{record.get('承銷商', '')}"
        )
    lines: list[str] = []
    for code in ("CI", "CB", "ECB", "GDR", "EB", "其他"):
        items = grouped.get(code, [])
        if not items:
            continue
        lines.append(f"{code}（{len(items)}件）")
        lines.extend(items)
        lines.append("")
    return lines


# ---------------------------------------------------------------------------
# Async job store — keeps report jobs in memory so Render's 30s HTTP timeout
# doesn't kill the generation mid-flight.
# ---------------------------------------------------------------------------
import threading as _threading
import uuid as _uuid

_JOBS: dict[str, dict[str, object]] = {}
_JOBS_LOCK = _threading.Lock()


def _create_job() -> str:
    job_id = _uuid.uuid4().hex
    with _JOBS_LOCK:
        _JOBS[job_id] = {"status": "running"}
    return job_id


def _finish_job(job_id: str, result: dict[str, object]) -> None:
    with _JOBS_LOCK:
        _JOBS[job_id] = {"status": "done", **result}


def _fail_job(job_id: str, error: str) -> None:
    with _JOBS_LOCK:
        _JOBS[job_id] = {"status": "error", "error": error}


def get_job(job_id: str) -> dict[str, object] | None:
    with _JOBS_LOCK:
        return _JOBS.get(job_id)


def week_from_records(records: list[dict[str, str]]) -> tuple[dt.date, dt.date]:
    """Derive Mon-Fri window from the latest 收文日期 in the source file."""
    latest: dt.date | None = None
    for r in records:
        d = parse_date(r.get("收文日期", ""))
        if d and (latest is None or d > latest):
            latest = d
    if latest is None:
        latest = current_friday()
    # Snap to the Friday of that week, then back to Monday
    monday = latest - dt.timedelta(days=latest.weekday())
    friday = monday + dt.timedelta(days=4)
    return monday, friday


def build_report(source_path: Path, source_url: str = "") -> dict[str, object]:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)
    records = extract_records(source_path)
    start, end = week_from_records(records)
    new_records = [r for r in records if in_range(r.get("收文日期", ""), start, end)]
    amend_records = [r for r in records if in_range(r.get("自動補正日期", ""), start, end)]
    stop_records = [r for r in records if in_range(r.get("停止生效日期", ""), start, end)]
    effective_records = [r for r in records if in_range(r.get("生效日期", ""), start, end)]
    weekly_records = unique_records(new_records + amend_records + stop_records + effective_records)
    # All CB/ECB/EB records need ordinal names for the 115年 summary sheet,
    # not just this week's new entries. CI/GDR only need company short names.
    all_bond_keys = {record_key(r) for r in records if r.get("分類") in ("CB", "ECB", "EB")}
    weekly_keys_set = {record_key(r) for r in weekly_records}
    lookup_focus = weekly_keys_set | all_bond_keys
    purpose_focus = weekly_keys_set
    lookup_warnings = enrich_records(records, end=end, focus_keys=lookup_focus, purpose_keys=purpose_focus)
    missing_purpose = [r for r in new_records if not r.get("本次籌資計畫", "").strip()]

    stats: dict[str, dict[str, int]] = {}
    for record in records:
        broker = record.get("承銷商", "") or "未填"
        code = record.get("分類", "其他")
        stats.setdefault(broker, {"合計": 0, "CI": 0, "CB": 0, "ECB": 0, "GDR": 0, "EB": 0, "其他": 0})
        stats[broker]["合計"] += 1
        stats[broker][code if code in stats[broker] else "其他"] += 1
    stat_rows = [["承銷商", "合計", "CI", "CB", "ECB", "GDR", "EB", "其他"]]
    for broker, values in sorted(stats.items(), key=lambda item: (-item[1]["合計"], item[0])):
        stat_rows.append([broker, values["合計"], values["CI"], values["CB"], values["ECB"], values["GDR"], values["EB"], values["其他"]])

    email_lines = [
        "Hi Everyone,",
        f"附件為 {roc_date(start)[:3]}/{start.month:02d}/{start.day:02d}～{roc_date(end)[:3]}/{end.month:02d}/{end.day:02d}同業送件明細，請查閱，謝謝！",
        f"本週新增{len(new_records)}家，生效{len(effective_records)}家",
        "",
        "以下為新增案件",
        *group_for_email(new_records),
        "以下為生效案件",
        *group_for_email(effective_records),
    ]

    summary_rows = [
        ["項目", "內容"],
        ["產出時間", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["週期", f"{roc_date(start)}～{roc_date(end)}"],
        ["來源檔", source_path.name],
        ["來源網址", source_url],
        ["篩選條件", "公司型態：上市、上櫃；案件類別：現金增資、轉換公司債、交換公司債、海外存託憑證"],
        ["本週新增", len(new_records)],
        ["本週自動補正", len(amend_records)],
        ["本週停止生效", len(stop_records)],
        ["本週生效", len(effective_records)],
        ["新增案件待補籌資原因", len(missing_purpose)],
    ]
    email_rows = [[line] for line in email_lines]
    counts = {
        "all": len(records),
        "new": len(new_records),
        "effective": len(effective_records),
        "amend": len(amend_records),
        "stop": len(stop_records),
        "missingPurpose": len(missing_purpose),
        "lookupWarnings": len(lookup_warnings),
    }

    template_path = TEMPLATE_PATH if TEMPLATE_PATH.exists() else None
    if template_path is not None:
        target = update_template_workbook_openpyxl(
            source_path,
            start,
            end,
            source_url,
            template_path,
            records,
            counts,
        )
    else:
        filename = f"同業送件明細_{compact_roc_date(start)}-{compact_roc_date(end)}.xlsx"
        target = REPORT_DIR / filename
        write_xlsx(
            target,
            [
                ("摘要", summary_rows),
                ("教學檔檢核", workflow_check_rows(start, end, source_path, source_url, counts)),
                ("本週更新總表", weekly_update_rows([
                    ("新增送件案件", "收文日期", new_records),
                    ("自動補正更新", "自動補正日期", amend_records),
                    ("停止生效更新", "停止生效日期", stop_records),
                    ("生效日期更新", "生效日期", effective_records),
                ])),
                ("本週新增", records_to_rows(new_records)),
                ("本週生效", records_to_rows(effective_records)),
                ("本週自動補正", records_to_rows(amend_records)),
                ("本週停止生效", records_to_rows(stop_records)),
                ("籌資目的檢核", purpose_check_rows(new_records)),
                ("承銷商統計", stat_rows),
                ("全部篩選資料", records_to_rows(records)),
                ("Email範本", email_rows),
            ],
        )
    email_path = target.with_suffix(".txt")
    email_path.write_text("\n".join(email_lines).strip() + "\n", encoding="utf-8")

    return {
        "file": target.name,
        "emailFile": email_path.name,
        "start": start.isoformat(),
        "end": end.isoformat(),
        "rocRange": f"{roc_date(start)}～{roc_date(end)}",
        "source": source_path.name,
        "sourceUrl": source_url,
        "counts": counts,
        "lookupWarnings": lookup_warnings,
        "email": "\n".join(email_lines).strip(),
    }


def list_reports() -> list[dict[str, object]]:
    ensure_dirs()
    files = sorted(REPORT_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return [
        {
            "file": p.name,
            "emailFile": p.with_suffix(".txt").name if p.with_suffix(".txt").exists() else "",
            "size": p.stat().st_size,
            "modified": dt.datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
        }
        for p in files
    ]



HTML = """<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>同業送件明細</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:"Segoe UI","Microsoft JhengHei",sans-serif;background:#f0f2f5;color:#1a2233;min-height:100vh}
    header{background:#12634f;color:#fff;padding:20px 24px}
    header h1{font-size:20px;font-weight:700}
    header p{font-size:13px;opacity:.75;margin-top:4px}
    .page{max-width:960px;margin:24px auto;padding:0 16px;display:grid;gap:16px}
    .card{background:#fff;border:1px solid #dde3ee;border-radius:10px;padding:20px}
    .card h2{font-size:15px;font-weight:700;color:#12634f;margin-bottom:16px}
    .upload-row{display:flex;gap:12px;align-items:flex-end;flex-wrap:wrap}
    .file-label{flex:1;min-width:200px}
    .file-label span{display:block;font-size:12px;color:#647084;margin-bottom:6px}
    .file-input{width:100%;padding:8px 10px;border:1.5px dashed #c2cfe0;border-radius:7px;font:inherit;color:#1a2233;cursor:pointer;background:#fafbfc}
    .file-input:hover{border-color:#12634f;background:#f0f7f4}
    .btn{padding:10px 22px;border:none;border-radius:7px;font:inherit;font-weight:700;font-size:14px;cursor:pointer;white-space:nowrap}
    .btn-primary{background:#12634f;color:#fff}
    .btn-primary:hover{background:#0e4f3f}
    .btn-primary:disabled{opacity:.55;cursor:not-allowed}
    .progress{margin-top:14px;display:none}
    .progress-bar-wrap{height:5px;background:#e8edf4;border-radius:3px;overflow:hidden;margin-bottom:8px}
    .progress-bar{height:100%;background:#12634f;border-radius:3px;width:0;transition:width .4s ease}
    .progress-text{font-size:13px;color:#647084}
    .status-msg{margin-top:10px;font-size:13px;line-height:1.6;white-space:pre-wrap;display:none}
    .status-msg.ok{color:#0a4a39}
    .status-msg.err{color:#c0001a}
    .metrics{display:none;grid-template-columns:repeat(auto-fit,minmax(120px,1fr));gap:10px;margin-top:14px}
    .metric{background:#f0f7f4;border-radius:8px;padding:12px}
    .metric strong{display:block;font-size:26px;font-weight:700;color:#12634f}
    .metric span{font-size:12px;color:#647084;margin-top:2px;display:block}
    textarea{width:100%;min-height:180px;border:1.5px solid #dde3ee;border-radius:7px;padding:10px;font:13px/1.6 Consolas,"Microsoft JhengHei",monospace;resize:vertical;color:#1a2233}
    .file-list{display:none}
    table{width:100%;border-collapse:collapse}
    th{font-size:12px;color:#647084;text-align:left;padding:6px 8px;border-bottom:1px solid #dde3ee}
    td{padding:8px;border-bottom:1px solid #f0f2f5;font-size:13px;vertical-align:middle}
    .dl-btn{display:inline-flex;padding:5px 12px;background:#e8f4f0;color:#12634f;border-radius:5px;font-size:12px;font-weight:700;text-decoration:none}
    .dl-btn:hover{background:#c9e8df}
    .empty-note{font-size:13px;color:#647084;padding:4px 0}
  </style>
</head>
<body>
<header>
  <h1>同業送件明細</h1>
  <p>上傳證期局申報案件彙總表，自動產出每週 Excel 與 Email 範本</p>
</header>
<div class="page">
  <div class="card">
    <h2>上傳來源檔</h2>
    <div class="upload-row">
      <div class="file-label">
        <span>證期局年度申報案件 Excel（.xlsx）</span>
        <input id="fileInput" class="file-input" type="file" accept=".xlsx,.xls">
      </div>
      <button id="genBtn" class="btn btn-primary" type="button">產出</button>
    </div>
    <div class="progress" id="progressWrap">
      <div class="progress-bar-wrap"><div class="progress-bar" id="progressBar"></div></div>
      <div class="progress-text" id="progressText">準備中…</div>
    </div>
    <div class="status-msg" id="statusMsg"></div>
    <div class="metrics" id="metrics">
      <div class="metric"><strong id="m-all">-</strong><span>篩選後總筆數</span></div>
      <div class="metric"><strong id="m-new">-</strong><span>本週新增</span></div>
      <div class="metric"><strong id="m-eff">-</strong><span>本週生效</span></div>
      <div class="metric"><strong id="m-amd">-</strong><span>補正／停止</span></div>
    </div>
  </div>
  <div class="card">
    <h2>Email 範本</h2>
    <textarea id="emailBox" readonly placeholder="產出完成後自動填入"></textarea>
  </div>
  <div class="card">
    <h2>已產出檔案</h2>
    <div class="empty-note" id="emptyNote">尚未產出，上傳後按「產出」即可。</div>
    <div class="file-list" id="fileList">
      <table><thead><tr><th>檔名</th><th>時間</th><th></th></tr></thead>
      <tbody id="fileRows"></tbody></table>
    </div>
  </div>
</div>
<script>
(function() {
  var fileInput    = document.getElementById("fileInput");
  var genBtn       = document.getElementById("genBtn");
  var progressWrap = document.getElementById("progressWrap");
  var progressBar  = document.getElementById("progressBar");
  var progressText = document.getElementById("progressText");
  var statusMsg    = document.getElementById("statusMsg");
  var metricsEl    = document.getElementById("metrics");
  var emailBox     = document.getElementById("emailBox");
  var emptyNote    = document.getElementById("emptyNote");
  var fileListEl   = document.getElementById("fileList");
  var fileRows     = document.getElementById("fileRows");

  var STEPS = [
    [5,  "上傳檔案中…"],
    [20, "篩選案件中…"],
    [40, "查詢公開資訊觀測站（CB/ECB 次數）…"],
    [65, "比對本週新增與生效…"],
    [80, "寫入 Excel 欄位中…"],
    [92, "產出郵件範本…"],
  ];
  var stepIdx = 0;
  var stepTimer = null;

  function startProgress() {
    stepIdx = 0;
    progressWrap.style.display = "block";
    progressBar.style.width = STEPS[0][0] + "%";
    progressText.textContent = STEPS[0][1];
    clearInterval(stepTimer);
    stepTimer = setInterval(function() {
      stepIdx = Math.min(stepIdx + 1, STEPS.length - 1);
      progressBar.style.width = STEPS[stepIdx][0] + "%";
      progressText.textContent = STEPS[stepIdx][1];
    }, 8000);
  }

  function stopProgress(ok) {
    clearInterval(stepTimer);
    stepTimer = null;
    progressBar.style.width = ok ? "100%" : "0%";
    progressText.textContent = ok ? "完成！" : "";
  }

  function showStatus(msg, isErr) {
    statusMsg.textContent = msg;
    statusMsg.className = "status-msg " + (isErr ? "err" : "ok");
    statusMsg.style.display = "block";
  }

  function showMetrics(c) {
    document.getElementById("m-all").textContent = c.all;
    document.getElementById("m-new").textContent = c.new;
    document.getElementById("m-eff").textContent = c.effective;
    document.getElementById("m-amd").textContent = (c.amend || 0) + "/" + (c.stop || 0);
    metricsEl.style.display = "grid";
  }

  function refreshFiles() {
    fetch("/api/reports")
      .then(function(r) { return r.json(); })
      .then(function(d) {
        var reps = d.reports || [];
        if (!reps.length) {
          emptyNote.style.display = "block";
          fileListEl.style.display = "none";
          return;
        }
        emptyNote.style.display = "none";
        fileListEl.style.display = "block";
        fileRows.innerHTML = reps.map(function(r) {
          return "<tr><td>" + r.file + "</td><td>" + r.modified +
            "</td><td><a class='dl-btn' href='/download/" +
            encodeURIComponent(r.file) + "'>下載</a></td></tr>";
        }).join("");
      })
      .catch(function() {});
  }

  genBtn.addEventListener("click", function() {
    var file = fileInput.files[0];
    if (!file) {
      showStatus("請先選擇 Excel 檔案。", true);
      return;
    }

    genBtn.disabled = true;
    metricsEl.style.display = "none";
    emailBox.value = "";
    statusMsg.style.display = "none";
    startProgress();

    var form = new FormData();
    form.append("source", file);

    fetch("/api/generate-upload", { method: "POST", body: form })
      .then(function(r) {
        return r.json().then(function(d) {
          if (!r.ok) { throw new Error(d.error || "上傳失敗 (" + r.status + ")"); }
          return d;
        });
      })
      .then(function(d) {
        var jobId = d.jobId;
        if (!jobId) { throw new Error("伺服器未回傳 jobId"); }

        var pollTimer = setInterval(function() {
          fetch("/api/job/" + jobId)
            .then(function(r) {
              return r.json().then(function(j) { return { ok: r.ok, j: j }; });
            })
            .then(function(x) {
              if (!x.ok) {
                clearInterval(pollTimer);
                stopProgress(false);
                showStatus("查詢失敗，請重新整理後再試。", true);
                genBtn.disabled = false;
                return;
              }
              var j = x.j;
              if (j.status === "done") {
                clearInterval(pollTimer);
                stopProgress(true);
                showMetrics(j.counts);
                emailBox.value = j.email || "";
                var warns = (j.lookupWarnings || []);
                var w = warns.length ? "\\n\\nMOPS 待確認：\\n" + warns.join("\\n") : "";
                showStatus("已產出：" + j.file + "\\n週期：" + j.rocRange + w, false);
                refreshFiles();
                genBtn.disabled = false;
              } else if (j.status === "error") {
                clearInterval(pollTimer);
                stopProgress(false);
                showStatus(j.error || "產出失敗", true);
                genBtn.disabled = false;
              }
            })
            .catch(function(e) {
              clearInterval(pollTimer);
              stopProgress(false);
              showStatus(e.message, true);
              genBtn.disabled = false;
            });
        }, 3000);
      })
      .catch(function(e) {
        stopProgress(false);
        showStatus(e.message, true);
        genBtn.disabled = false;
      });
  });
})();
</script>
</body>
</html>
"""



class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt: str, *args: object) -> None:
        sys.stderr.write("[%s] %s\n" % (self.log_date_time_string(), fmt % args))

    def send_json(self, status: int, payload: dict[str, object]) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_HEAD(self) -> None:
        self.send_response(200)
        self.end_headers()

    def do_GET(self) -> None:
        ensure_dirs()
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/healthz":
            self.send_json(200, {"ok": True})
            return
        if parsed.path == "/":
            body = HTML.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if parsed.path == "/api/reports":
            self.send_json(200, {"reports": list_reports()})
            return
        if parsed.path.startswith("/api/job/"):
            job_id = parsed.path.removeprefix("/api/job/")
            job = get_job(job_id)
            if job is None:
                self.send_json(404, {"error": "job not found"})
            else:
                self.send_json(200, job)
            return
        if parsed.path == "/api/diagnostics":
            self.send_json(200, {"results": run_download_diagnostics()})
            return
        if parsed.path.startswith("/download/"):
            name = urllib.parse.unquote(parsed.path.removeprefix("/download/"))
            safe_name = Path(name).name
            path = REPORT_DIR / safe_name
            if not path.exists():
                self.send_error(404, "file not found")
                return
            body = path.read_bytes()
            self.send_response(200)
            content_type = "text/plain; charset=utf-8" if path.suffix.lower() == ".txt" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{urllib.parse.quote(safe_name)}")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        self.send_error(404)

    def do_POST(self) -> None:
        ensure_dirs()
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/api/generate-upload":
            self.handle_upload_generate(parsed)
            return
        if parsed.path != "/api/generate":
            self.send_error(404)
            return
        try:
            source_path, source_url = download_latest_source(current_friday())
            result = build_report(source_path, source_url)
            self.send_json(200, result)
        except Exception as exc:
            self.send_json(500, {"error": str(exc)})

    def handle_upload_generate(self, parsed: urllib.parse.ParseResult) -> None:
        try:
            source_path, fields = self.save_uploaded_xlsx()
        except Exception as exc:
            self.send_json(400, {"error": str(exc)})
            return
        job_id = _create_job()

        def _run() -> None:
            try:
                result = build_report(source_path, "uploaded")
                _finish_job(job_id, result)
            except Exception as exc:
                _fail_job(job_id, str(exc))

        _threading.Timer(600, lambda: _fail_job(job_id, "產出超時（10 分鐘），MOPS 查詢可能無回應，請重試。") if get_job(job_id) and get_job(job_id).get("status") == "running" else None).start()

        _threading.Thread(target=_run, daemon=True).start()
        self.send_json(202, {"jobId": job_id})

    def save_uploaded_xlsx(self) -> tuple[Path, dict[str, str]]:
        content_type = self.headers.get("Content-Type", "")
        match = re.search(r"boundary=(.+)", content_type)
        if not match:
            raise ValueError("找不到上傳檔案。")
        boundary = match.group(1).strip().strip('"').encode()
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length)
        parts = body.split(b"--" + boundary)
        fields: dict[str, str] = {}
        source_path: Path | None = None
        for part in parts:
            if b"Content-Disposition:" not in part:
                continue
            header, _, data = part.partition(b"\r\n\r\n")
            if not data:
                continue
            name_match = re.search(rb'name="([^"]*)"', header)
            field_name = name_match.group(1).decode("utf-8", errors="ignore") if name_match else ""
            data = data.rstrip(b"\r\n")
            if data.endswith(b"--"):
                data = data[:-2].rstrip(b"\r\n")
            if b"filename=" not in header:
                if field_name:
                    fields[field_name] = data.decode("utf-8", errors="ignore").strip()
                continue
            filename_match = re.search(rb'filename="([^"]*)"', header)
            filename = filename_match.group(1).decode("utf-8", errors="ignore") if filename_match else "source.xlsx"
            filename = Path(filename).name or "source.xlsx"
            if not filename.lower().endswith((".xlsx", ".xls")):
                raise ValueError("請上傳 Excel 檔。")
            target = SOURCE_DIR / f"source_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
            target.write_bytes(data)
            source_path = target
        if source_path is not None:
            return source_path, fields
        raise ValueError("沒有收到證期局來源 Excel 檔。")


def main() -> None:
    ensure_dirs()
    port = int(os.environ.get("PORT", "8787"))
    server = ThreadingHTTPServer(("0.0.0.0", port), Handler)
    print(f"同業送件明細下載網頁已啟動：http://localhost:{port}")
    print("同公司網路使用者可用這台電腦的 IP 加上同一個 port 連線。")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n已停止。")


if __name__ == "__main__":
    main()
